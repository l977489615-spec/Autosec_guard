#!/usr/bin/env python3
"""Audit paper docx + workbook against frozen strict experiment data."""

from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from l3_evidence_rates import (
    ablation_evidence_rates,
    model_evidence_rates,
    pentestgpt_evidence_rate,
    scan_evidence_by_category,
    table10_evidence_rates,
)
from paper_metric_names import (
    EVIDENCE_COMPLETENESS_COL,
    JSON_LATENCY_KEY,
    JSON_RECALL_KEY,
    JSON_SUBTASK_KEY,
    MISS_RATE_COL,
    RECALL_AT_GT_COL,
    SUBTASK_COMPLETION_COL,
)
from paper_table_rows import load_json
from update_paper_v3_0610_data_only import pct_only


ROOT = Path(__file__).resolve().parents[1]
PAPER_DOCX = Path(
    "/Users/queen/Desktop/ICV_POC_research/论文/"
    "面向智能网联汽车的证据驱动多智能体协同漏洞验证方法研究_6.9_1_贡献与章节结构修订版_v3_0610.docx"
)
WORKBOOK = Path("/Users/queen/Desktop/ICV_POC_research/论文/论文内表格（新）.xlsx")


def visible_docx_text(docx_path: Path) -> str:
    with zipfile.ZipFile(docx_path) as package:
        xml = package.read("word/document.xml").decode("utf-8")
    xml = re.sub(r"<w:del[^>]*>.*?</w:del>", "", xml, flags=re.S)
    return "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, re.S))


def workbook_row(sheet_name: str, key_col_value: str) -> dict[str, object]:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[sheet_name]
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == key_col_value:
            return {header: ws.cell(row, col).value for header, col in headers.items()}
    raise KeyError(f"{sheet_name}: 未找到行 {key_col_value!r}")


def expect_equal(issues: list[str], label: str, actual: object, expected: object) -> None:
    if actual != expected:
        issues.append(f"{label}: 实际={actual!r}，期望={expected!r}")


def audit_workbook(issues: list[str]) -> None:
    table6 = next(row for row in load_json("table6_total_by_category.json") if row["类别"] == "合计")
    table7 = {row["组别"]: row for row in load_json("table7_total_by_model_group.json")}
    table8 = {row["variant_id"]: row for row in load_json("table8_total_by_model.json")}
    ab_ev = ablation_evidence_rates()
    mod_ev = model_evidence_rates()
    t10_ev = table10_evidence_rates()

    row = workbook_row("表6_PoC命中成功漏报（智谱）", "合计")
    for key, expected in [
        (RECALL_AT_GT_COL, table6[JSON_RECALL_KEY]),
        (SUBTASK_COMPLETION_COL, table6[JSON_SUBTASK_KEY]),
        (MISS_RATE_COL, table6["漏报率"]),
    ]:
        expect_equal(issues, f"Excel表6合计/{key}", row.get(key), expected)

    for group in "ABCD":
        row = workbook_row("表7_智能体消融（智谱）", group)
        src = table7[group]
        expect_equal(issues, f"Excel表7{group}/召回", row.get(RECALL_AT_GT_COL), src[JSON_RECALL_KEY])
        expect_equal(issues, f"Excel表7{group}/子任务", row.get(SUBTASK_COMPLETION_COL), src[JSON_SUBTASK_KEY])
        expect_equal(issues, f"Excel表7{group}/证据", row.get(EVIDENCE_COMPLETENESS_COL), ab_ev[group])

    zhipu_row = workbook_row("表8_模型对比", "智谱 GLM-5")
    zhipu = table8["ZHIPU"]
    expect_equal(issues, "Excel表8智谱/召回", zhipu_row.get(RECALL_AT_GT_COL), zhipu[JSON_RECALL_KEY])
    expect_equal(issues, "Excel表8智谱/子任务", zhipu_row.get(SUBTASK_COMPLETION_COL), zhipu[JSON_SUBTASK_KEY])
    expect_equal(issues, "Excel表8智谱/证据", zhipu_row.get(EVIDENCE_COMPLETENESS_COL), mod_ev["ZHIPU"])
    expect_equal(issues, "Excel表8智谱/耗时", zhipu_row.get(JSON_LATENCY_KEY), zhipu[JSON_LATENCY_KEY])

    deepseek_row = workbook_row("表8_模型对比", "DeepSeek v4 pro")
    deepseek = table8["DEEPSEEK"]
    expect_equal(
        issues,
        "Excel表8 DeepSeek/子任务",
        deepseek_row.get(SUBTASK_COMPLETION_COL),
        deepseek[JSON_SUBTASK_KEY],
    )

    for label, expected in t10_ev.items():
        row = workbook_row("表10_平台能力对比", label)
        expect_equal(issues, f"Excel表10/{label}/证据", row.get(EVIDENCE_COMPLETENESS_COL), expected)


def audit_docx_prose(issues: list[str], plain: str) -> None:
    table6 = next(row for row in load_json("table6_total_by_category.json") if row["类别"] == "合计")
    table7d = next(row for row in load_json("table7_total_by_model_group.json") if row["组别"] == "D")
    table8 = {row["variant_id"]: row for row in load_json("table8_total_by_model.json")}

    required_snippets = [
        pct_only(table6[JSON_RECALL_KEY]),
        pct_only(table6[JSON_SUBTASK_KEY]),
        pct_only(table6["漏报率"]),
        pct_only(model_evidence_rates()["ZHIPU"]),
        table8["ZHIPU"][JSON_LATENCY_KEY],
        scan_evidence_by_category()["合计"],
        "46.7%",
        "58.7%",
        "13.3%",
    ]
    for snippet in required_snippets:
        if snippet not in plain:
            issues.append(f"正文缺少关键数值/表述: {snippet}")

    banned = [
        "任务推进率",
        "可审计证据率",
        "宏平均成功率",
        "漏洞检出率",
        "有效证据率",
        "平均时延",
        "97.7%",
        "95.8%",
        "53.3%（16/30）",
    ]
    for token in banned:
        if token in plain:
            issues.append(f"正文含应删除/已废弃内容: {token}")

    if "DeepSeek v4 pro与智谱GLM-5的基准子任务完成率均达到93.3%" in plain:
        issues.append("§6.5 模型适配段仍将 DeepSeek 子任务完成率误写为 93.3%")

    if "实验采用基准阳性召回率、基准子任务完成率、证据完整率和平均端到端净耗时五项指标评价" in plain:
        issues.append("中文摘要五项指标枚举缺少漏报率")

    if (
        "Miss Rate, Evidence Completeness Rate, and Mean End-to-End Runtime. EDVV achieves"
        in plain
    ):
        issues.append("英文摘要五项指标枚举缺少 Miss Rate")

    if "已完成验证项数/基准阳性PoC数" in plain:
        issues.append("§6.3 子任务完成率公式分母表述错误")

    if pct_only(table7d[JSON_SUBTASK_KEY]) not in plain:
        issues.append("正文未出现 EDVV（智谱）子任务完成率 93.3%")


def main() -> int:
    issues: list[str] = []
    if not WORKBOOK.is_file():
        issues.append(f"未找到工作簿: {WORKBOOK}")
    else:
        audit_workbook(issues)

    if not PAPER_DOCX.is_file():
        issues.append(f"未找到论文: {PAPER_DOCX}")
    else:
        audit_docx_prose(issues, visible_docx_text(PAPER_DOCX))

    report = {"ok": not issues, "issue_count": len(issues), "issues": issues}
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if not issues else 1


if __name__ == "__main__":
    sys.path.insert(0, str(ROOT / "lab"))
    raise SystemExit(main())
