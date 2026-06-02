#!/usr/bin/env python3
import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def read_json(path: Path, default: Any):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for index, header in enumerate(headers, start=1):
        width = max(12, len(str(header)) + 2)
        for row_index in range(2, min(ws.max_row, 80) + 1):
            width = max(width, min(45, len(str(ws.cell(row_index, index).value or "")) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    return ws


def stringify(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        return ",".join(map(str, value))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build advisor-required experiment workbook.")
    parser.add_argument("--experiment-dir", type=Path, default=Path("lab/evidence"))
    parser.add_argument("--can-records", type=Path, default=Path("lab/can_test_records.csv"))
    parser.add_argument("--output", type=Path, default=Path("lab/实验数据统计表.xlsx"))
    args = parser.parse_args()

    coverage = read_json(args.experiment_dir / "poc_coverage.json", {"pocs": [], "by_category": {}, "by_attack_surface": {}, "total": 0})
    scan_results = read_json(args.experiment_dir / "scan_results.json", [])
    agent_rows = read_json(args.experiment_dir / "agent_orchestration.json", [])
    edge_rows = read_json(args.experiment_dir / "edge_capabilities.json", [])
    comparison_rows = read_json(args.experiment_dir / "comparison.json", [])
    typical_cases = read_json(args.experiment_dir / "typical_cases.json", [])
    can_rows = read_csv(args.can_records)

    wb = Workbook()
    wb.remove(wb.active)

    coverage_rows = []
    for name, count in sorted((coverage.get("by_category") or {}).items()):
        coverage_rows.append({"统计维度": "PoC分类", "名称": name, "数量": count, "用途": "证明平台覆盖面"})
    for name, count in sorted((coverage.get("by_attack_surface") or {}).items()):
        coverage_rows.append({"统计维度": "攻击面", "名称": name, "数量": count, "用途": "证明攻击面覆盖"})
    coverage_rows.append({"统计维度": "总计", "名称": "PoC总数", "数量": coverage.get("total", 0), "用途": "证明插件库规模"})
    write_sheet(wb, "表1_PoC覆盖情况", ["统计维度", "名称", "数量", "用途"], coverage_rows)

    poc_rows = [{
        "PoC编号": row.get("display_id") or row.get("poc_file"),
        "PoC文件": row.get("poc_file"),
        "检测项": row.get("poc_name"),
        "分类": row.get("category"),
        "攻击面": row.get("attack_surface"),
        "协议": row.get("protocol"),
        "风险等级": row.get("severity"),
        "破坏等级": row.get("destructive_level"),
        "是否高危": row.get("high_risk"),
        "自动Profile": stringify(row.get("profiles")),
        "必需参数": row.get("required_params"),
    } for row in coverage.get("pocs", [])]
    write_sheet(wb, "PoC插件分类清单", ["PoC编号", "PoC文件", "检测项", "分类", "攻击面", "协议", "风险等级", "破坏等级", "是否高危", "自动Profile", "必需参数"], poc_rows)

    scan_sheet_rows = [{
        "目标": row.get("target_id"),
        "PoC编号": row.get("poc_display_id") or row.get("poc_file"),
        "检测项": row.get("poc_name"),
        "类别": row.get("category"),
        "攻击面": row.get("attack_surface"),
        "自动Profile": row.get("auto_profile"),
        "执行状态": row.get("status"),
        "耗时s": row.get("elapsed_seconds"),
        "是否发现风险": row.get("vulnerable"),
        "是否拦截": row.get("blocked"),
        "证据文件": row.get("evidence_file"),
        "证据摘要": row.get("evidence"),
    } for row in scan_results]
    write_sheet(wb, "表2_扫描执行结果", ["目标", "PoC编号", "检测项", "类别", "攻击面", "自动Profile", "执行状态", "耗时s", "是否发现风险", "是否拦截", "证据文件", "证据摘要"], scan_sheet_rows)

    agent_sheet_rows = [{
        "任务ID": row.get("task_id"),
        "目标IP": row.get("target_ip"),
        "方法": row.get("method"),
        "规划PoC数量": row.get("planned_poc_count"),
        "调用PoC数量": row.get("executed_poc_count"),
        "反思重入次数": row.get("reflection_reentry_count"),
        "失败/重试计数": row.get("retry_or_error_count"),
        "发现数量": row.get("finding_count"),
        "耗时s": row.get("elapsed_seconds"),
        "报告文件": row.get("report_file"),
    } for row in agent_rows]
    write_sheet(wb, "表3_多Agent编排", ["任务ID", "目标IP", "方法", "规划PoC数量", "调用PoC数量", "反思重入次数", "失败/重试计数", "发现数量", "耗时s", "报告文件"], agent_sheet_rows)

    write_sheet(wb, "表4_CAN网关联动", list(can_rows[0].keys()) if can_rows else ["case_id", "test_type", "interface", "can_id", "frame_type", "payload_hex", "send_count", "observed_response", "abnormal", "blocked_by_safety", "evidence_file"], can_rows)

    safety_rows = [{
        "PoC编号": row.get("display_id") or row.get("poc_file"),
        "检测项": row.get("poc_name"),
        "破坏等级": row.get("destructive_level"),
        "是否扰动": row.get("is_disruptive"),
        "是否高危": row.get("high_risk"),
        "必需参数": row.get("required_params"),
    } for row in coverage.get("pocs", []) if row.get("high_risk")]
    blocked_rows = [{
        "PoC编号": row.get("poc_display_id") or row.get("poc_file"),
        "目标": row.get("target_id"),
        "执行状态": row.get("status"),
        "是否需要授权": row.get("requires_approval"),
        "证据文件": row.get("evidence_file"),
    } for row in scan_results if row.get("blocked") or row.get("requires_approval")]
    write_sheet(wb, "安全控制记录", ["PoC编号", "检测项", "破坏等级", "是否扰动", "是否高危", "必需参数"], safety_rows)
    write_sheet(wb, "安全拦截明细", ["PoC编号", "目标", "执行状态", "是否需要授权", "证据文件"], blocked_rows)

    edge_sheet_rows = [{k: stringify(v) for k, v in row.items()} for row in edge_rows]
    write_sheet(wb, "边缘设备能力", list(edge_sheet_rows[0].keys()) if edge_sheet_rows else ["target_id", "target_ip", "can_interface", "bluetooth_mac", "wifi_interface"], edge_sheet_rows)

    comparison_sheet_rows = [{k: stringify(v) for k, v in row.items()} for row in comparison_rows]
    write_sheet(wb, "对比实验数据", list(comparison_sheet_rows[0].keys()) if comparison_sheet_rows else ["case_id", "target_id", "manual_elapsed_seconds", "platform_elapsed_seconds"], comparison_sheet_rows)

    evidence_rows = [{
        "目标": row.get("target_id"),
        "PoC编号": row.get("poc_display_id") or row.get("poc_file"),
        "执行参数/结果JSON": row.get("evidence_file"),
        "证据摘要": row.get("evidence"),
    } for row in scan_results]
    evidence_rows.extend({
        "目标": row.get("task_id"),
        "PoC编号": "Agent workflow",
        "执行参数/结果JSON": row.get("report_file"),
        "证据摘要": f"planned={row.get('planned_poc_count')}, executed={row.get('executed_poc_count')}, findings={row.get('finding_count')}",
    } for row in agent_rows)
    write_sheet(wb, "证据归档数据", ["目标", "PoC编号", "执行参数/结果JSON", "证据摘要"], evidence_rows)

    case_rows = [{k: stringify(v) for k, v in row.items()} for row in typical_cases]
    write_sheet(wb, "典型案例数据", list(case_rows[0].keys()) if case_rows else ["case_id", "case_name", "target_id", "attack_surface", "expected_evidence"], case_rows)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(json.dumps({"output": str(args.output)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
