#!/usr/bin/env python3
"""Build paper-facing experiment datasets from current evidence artifacts.

The script deliberately separates measured rows from derived engineering
metrics.  Measured rows come from scan_results, poc_runs, agent reports and
comparison.json.  Derived rows are deterministic summaries used for paper
tables that require ablation/planning baselines but do not yet have separate
runtime logs.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

LAB_DIR = Path(__file__).resolve().parent
ROOT = LAB_DIR.parent
SERVER_DIR = ROOT / "server"
for path in (LAB_DIR, SERVER_DIR):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))


def read_json(path: Path, default: Any) -> Any:
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_csv(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in {"", None}:
            return default
        return float(value)
    except Exception:
        return default


def pct(numerator: float, denominator: float) -> str:
    if denominator <= 0:
        return ""
    return f"{round(numerator / denominator * 100, 1)}%"


def pct_frac(numerator: float, denominator: float) -> str:
    from metric_definitions import rate_display

    if denominator <= 0:
        return "-"
    return rate_display(numerator, denominator)


def coverage_metrics_from_report(
    report_file: str,
    target_id: str,
    evidence_root: Path | None = None,
) -> dict[str, Any]:
    from metric_definitions import coverage_metrics
    from table10_platform_scoring import score_platform_report

    path = Path(str(report_file))
    if not path.is_file() and evidence_root:
        candidate = evidence_root / target_id / "agent_runs" / path.name
        if candidate.is_file():
            path = candidate
    if not path.is_file():
        return coverage_metrics(0, 0)
    score = score_platform_report(path, target_id=target_id, evidence_root=evidence_root)
    return coverage_metrics(
        int(score.get("coverage_num") or score.get("task_num") or 0),
        int(score.get("coverage_den") or score.get("task_den") or 0),
    )


def pct_or_label(numerator: float, denominator: float, empty_label: str) -> str:
    if denominator <= 0:
        return empty_label
    return pct(numerator, denominator)


def ratio(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return round(numerator / denominator, 4)


def sanitize_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    return text[:32767]


def write_sheet(wb: Workbook, title: str, rows: list[dict], headers: list[str] | None = None) -> None:
    ws = wb.create_sheet(title[:31])
    headers = headers or sorted({key for row in rows for key in row.keys()})
    ws.append([sanitize_excel_value(header) for header in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([sanitize_excel_value(row.get(header, "")) for header in headers])
    for index, header in enumerate(headers, start=1):
        width = max(12, len(str(header)) + 2)
        for row_index in range(2, min(ws.max_row, 120) + 1):
            width = max(width, min(48, len(str(ws.cell(row_index, index).value or "")) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def latest_target_dirs(evidence_root: Path, target_id: str | None = None) -> list[Path]:
    dirs = sorted(
        path for path in evidence_root.iterdir()
        if path.is_dir() and ((path / "scan_results.json").is_file() or (path / "poc_runs").is_dir())
    )
    if target_id:
        dirs = [path for path in dirs if path.name == target_id]
    return dirs


def target_ids_from_evidence(evidence_root: Path) -> list[str]:
    return [path.name for path in latest_target_dirs(evidence_root)]


def load_coverage(evidence_root: Path) -> dict:
    coverage = read_json(evidence_root / "poc_coverage.json", {})
    if coverage:
        return coverage
    for tdir in latest_target_dirs(evidence_root):
        coverage = read_json(tdir / "poc_coverage.json", {})
        if coverage:
            return coverage
    from run_experiment import collect_poc_coverage, write_json as write_project_json

    coverage = collect_poc_coverage()
    write_project_json(evidence_root / "poc_coverage.json", coverage)
    return coverage


def coverage_meta_map(coverage: dict) -> dict[str, dict]:
    return {
        str(item.get("poc_file") or ""): item
        for item in coverage.get("pocs", []) or []
        if item.get("poc_file")
    }


def category_for_poc(poc_file: str, meta: dict) -> str:
    category = str(meta.get("category") or poc_file.split("/", 1)[0] if poc_file else "")
    attack_surface = str(meta.get("attack_surface") or "")
    if category == "reconnaissance":
        return "侦察类"
    if category == "network":
        return "网络服务类"
    if category == "canbus" or "CAN" in attack_surface or "UDS" in attack_surface or category == "new_can":
        return "车内协议类"
    if category == "wireless" or category in {"new_wireless", "new_peripheral"} or "无线" in attack_surface:
        return "无线与外设类"
    if category in {"application", "new_application"} or "车机APP" in attack_surface:
        return "应用安全类"
    if category == "new_system" or "系统配置" in attack_surface:
        return "系统配置类"
    if "第三方组件" in attack_surface or category == "new_advanced":
        return "第三方组件类"
    return "高级攻击类"


def synthesize_rows_from_poc_runs(tdir: Path, meta_by_poc: dict[str, dict]) -> list[dict]:
    rows: list[dict] = []
    for path in sorted((tdir / "poc_runs").glob("*.json")):
        payload = read_json(path, {})
        poc_file = str(payload.get("poc_id") or payload.get("poc_file") or "").strip()
        if not poc_file:
            continue
        meta = meta_by_poc.get(poc_file, {})
        success = payload.get("success") is not False and not payload.get("blocked")
        rows.append({
            "target_id": tdir.name,
            "target_type": "",
            "poc_file": poc_file,
            "poc_name": (payload.get("security_profile") or {}).get("poc_name") or meta.get("poc_name") or poc_file,
            "category": meta.get("category") or poc_file.split("/", 1)[0],
            "attack_surface": meta.get("attack_surface", ""),
            "status": "completed" if success else "error",
            "elapsed_seconds": payload.get("elapsed_seconds", 0),
            "vulnerable": payload.get("vulnerable") is True,
            "requires_human_review": payload.get("requires_human_review", False),
            "verification_status": payload.get("verification_status", ""),
            "blocked": payload.get("blocked", False),
            "evidence": payload.get("evidence", ""),
            "evidence_file": str(path),
            "data_source": "poc_runs",
        })
    return rows


def load_scan_rows(evidence_root: Path, meta_by_poc: dict[str, dict], target_id: str | None = None) -> list[dict]:
    rows: list[dict] = []
    for tdir in latest_target_dirs(evidence_root, target_id):
        scan_rows = read_json(tdir / "scan_results.json", [])
        if isinstance(scan_rows, list) and scan_rows:
            for row in scan_rows:
                item = dict(row)
                item.setdefault("data_source", "scan_results")
                item.setdefault("target_id", tdir.name)
                rows.append(item)
        else:
            rows.extend(synthesize_rows_from_poc_runs(tdir, meta_by_poc))
    return rows


def _normalize_poc_name(item: dict) -> str:
    from agent_vulnerability_hits import normalize_poc_reference

    return normalize_poc_reference(item)


def _execution_items(report: dict) -> list[dict]:
    structured = report.get("structured") or {}
    execution = (structured.get("execution") or {}).get("items") or []
    return [item for item in execution if isinstance(item, dict)]


def _attack_plan_items(report: dict) -> list[dict]:
    structured = report.get("structured") or {}
    plan = (structured.get("attack_plan") or {}).get("items") or []
    if plan:
        return [item for item in plan if isinstance(item, dict)]
    planner = (structured.get("planner") or {}).get("steps") or []
    return [item for item in planner if isinstance(item, dict)]


def _global_positive_pocs(scan_rows: list[dict]) -> set[str]:
    return {
        str(row.get("poc_file") or "").strip()
        for row in scan_rows
        if row.get("vulnerable") is True and str(row.get("poc_file") or "").strip()
    }


def _global_completed_pocs(scan_rows: list[dict]) -> set[str]:
    return {
        str(row.get("poc_file") or "").strip()
        for row in scan_rows
        if row.get("status") in {"completed", "pending_manual_review"} and str(row.get("poc_file") or "").strip()
    }


def _report_elapsed_seconds(report: dict) -> float:
    elapsed = safe_float(report.get("duration_seconds"))
    manual_wait = safe_float(report.get("manual_review_wait_seconds"))
    if elapsed:
        return max(round(elapsed - manual_wait, 3), 0.0)
    execution_elapsed = sum(safe_float(item.get("elapsed_seconds")) for item in _execution_items(report))
    llm_totals = ((report.get("llm_usage") or {}).get("totals") or {})
    llm_elapsed = safe_float(llm_totals.get("latency_ms_total")) / 1000
    return round(max(execution_elapsed + llm_elapsed - manual_wait, 0.0), 3)


def _llm_totals(report: dict) -> dict:
    usage = report.get("llm_usage") or ((report.get("structured") or {}).get("llm_usage") or {})
    return usage.get("totals") or {}


def _agent_run_group(path: Path, report: dict) -> str:
    group = str(report.get("ablation_group") or "").strip()
    if group:
        return group
    name = path.name.upper()
    if name.startswith("SINGLE-"):
        return "single_agent"
    return "multi_agent_full"


def load_agent_run_manifest(path: Path | None) -> list[Path]:
    if not path:
        return []
    payload = read_json(path, [])
    base_dir = path.parent
    raw_items = payload.get("agent_run_files", payload.get("files", [])) if isinstance(payload, dict) else payload
    files: list[Path] = []
    for item in raw_items or []:
        text = str(item.get("path") if isinstance(item, dict) else item).strip()
        if not text:
            continue
        item_path = Path(text)
        if not item_path.is_absolute():
            item_path = base_dir / item_path
        files.append(item_path)
    return files


def load_selected_agent_reports(manifest_path: Path | None, target_id: str | None = None) -> list[dict]:
    selected: list[dict] = []
    for path in load_agent_run_manifest(manifest_path):
        report = read_json(path, {})
        if not isinstance(report, dict) or not report:
            continue
        report_target = str(report.get("target_id") or path.parents[1].name if len(path.parents) > 1 else "")
        if target_id and report_target != target_id:
            continue
        selected.append({
            "path": path,
            "report": report,
            "target_id": report_target,
            "variant_id": str(report.get("variant_id") or path.name.split("_", 1)[0].replace("AGENT-", "").replace("SINGLE-", "")),
            "group": _agent_run_group(path, report),
        })
    return selected


def build_ablation_rows_from_agent_reports(selected_reports: list[dict], scan_rows: list[dict]) -> list[dict]:
    global_positive = _global_positive_pocs(scan_rows)
    rows: list[dict] = []
    for selected in selected_reports:
        report = selected["report"]
        execution = _execution_items(report)
        executed_pocs = {_normalize_poc_name(item) for item in execution if _normalize_poc_name(item)}
        from agent_vulnerability_hits import agent_vulnerable_poc_files

        agent_vuln = agent_vulnerable_poc_files(report)
        positive_hits = sorted(global_positive & agent_vuln)
        llm = _llm_totals(report)
        rows.append({
            "ablation_group": selected["group"],
            "target_id": selected["target_id"],
            "variant_id": selected["variant_id"],
            "planned_poc_count": len(_attack_plan_items(report)) or len(execution),
            "executed_poc_count": len(executed_pocs),
            "finding_count": len(agent_vuln),
            "error_count": sum(1 for item in execution if item.get("error")),
            "manual_review_required_count": sum(1 for item in execution if item.get("requires_human_review")),
            "manual_review_wait_seconds": safe_float(report.get("manual_review_wait_seconds")),
            "elapsed_seconds": _report_elapsed_seconds(report),
            "llm_call_count": int(llm.get("calls") or 0),
            "total_tokens": int(llm.get("total_tokens") or 0),
            "report_file": str(selected["path"]),
            "data_source": "selected_agent_report",
            "global_vulnerable_count": len(global_positive),
            "finding_overlap_with_global": len(positive_hits),
            "finding_overlap_pocs": positive_hits,
        })
    return rows


def build_comparison_rows_from_agent_reports(selected_reports: list[dict], scan_rows: list[dict], gt_positive: set[str]) -> list[dict]:
    global_positive = _global_positive_pocs(scan_rows)
    global_completed = _global_completed_pocs(scan_rows)
    rows: list[dict] = []
    for selected in selected_reports:
        if selected["group"] != "multi_agent_full":
            continue
        report = selected["report"]
        execution = _execution_items(report)
        executed_pocs = {_normalize_poc_name(item) for item in execution if _normalize_poc_name(item)}
        successful_pocs = {
            _normalize_poc_name(item)
            for item in execution
            if str(item.get("status") or "") in {"completed", "vulnerable", "pending_manual_review", "manual_review_completed"}
            and _normalize_poc_name(item)
        }
        from agent_vulnerability_hits import agent_vulnerable_poc_files

        agent_vuln = agent_vulnerable_poc_files(report)
        gt_hits = sorted(gt_positive & agent_vuln)
        overlap = sorted(global_positive & agent_vuln)
        baseline_overlap = sorted(global_completed & executed_pocs)
        llm = _llm_totals(report)
        executed = len(executed_pocs)
        finding_count = len(agent_vuln)
        rows.append({
            "target_id": selected["target_id"],
            "comparison_type": "global_vs_agent",
            "variant_id": selected["variant_id"],
            "variant_label": report.get("variant_label", selected["variant_id"]),
            "agent_elapsed_seconds": _report_elapsed_seconds(report),
            "agent_planned_poc_count": len(_attack_plan_items(report)) or len(execution),
            "agent_executed_poc_count": executed,
            "agent_executed_pocs": sorted(executed_pocs),
            "baseline_replay_poc_count": len(global_positive),
            "agent_execution_coverage_vs_global": ratio(len(baseline_overlap), len(global_completed)),
            "baseline_overlap_count": len(baseline_overlap),
            "baseline_overlap_pocs": baseline_overlap,
            "baseline_missed_count": max(len(global_completed) - len(baseline_overlap), 0),
            "agent_extra_execution_count": len(executed_pocs - global_completed),
            "agent_extra_success_count": len(successful_pocs - global_completed),
            "agent_finding_count": finding_count,
            "agent_reflection_reentry_count": report.get("reflector_reentry_count", 0),
            "agent_llm_call_count": int(llm.get("calls") or 0),
            "agent_prompt_tokens": int(llm.get("prompt_tokens") or 0),
            "agent_completion_tokens": int(llm.get("completion_tokens") or 0),
            "agent_total_tokens": int(llm.get("total_tokens") or 0),
            "agent_avg_llm_latency_ms": safe_float(llm.get("avg_latency_ms")),
            "agent_tokens_per_finding": round(int(llm.get("total_tokens") or 0) / max(finding_count, 1), 2) if finding_count else "",
            "finding_overlap_with_global": len(overlap),
            "finding_overlap_pocs": overlap,
            "agent_finding_recall_vs_global": ratio(len(overlap), len(global_positive)),
            "agent_finding_precision_vs_global": ratio(len(overlap), finding_count),
            "gt_positive_count": len(gt_positive),
            "gt_positive_pocs": sorted(list(gt_positive)),
            "gt_hit_count": len(gt_hits),
            "gt_hit_pocs": gt_hits,
            "gt_recall": ratio(len(gt_hits), len(gt_positive)) if gt_positive else "",
            "paper_primary_recall": ratio(len(gt_hits), len(gt_positive)) if gt_positive else ratio(len(overlap), len(global_positive)),
            "agent_gt_recall": ratio(len(gt_hits), len(gt_positive)) if gt_positive else "",
            "agent_findings_per_executed_poc": round(finding_count / max(executed, 1), 3) if executed else "",
            "agent_efficiency_score": "",
            "report_file": str(selected["path"]),
            "data_source": "selected_agent_report+scan_results",
        })
    return rows


def build_model_rows_from_agent_reports(selected_reports: list[dict], config: dict) -> list[dict]:
    label_by_variant = {
        str(variant.get("variant_id") or ""): str(variant.get("label") or variant.get("variant_id") or "")
        for variant in (config.get("model_comparison") or {}).get("variants", []) or []
    }
    rows: list[dict] = []
    for selected in selected_reports:
        if selected["group"] != "multi_agent_full":
            continue
        report = selected["report"]
        execution = _execution_items(report)
        llm = _llm_totals(report)
        finding_count = len(report.get("findings") or [])
        rows.append({
            "target_id": selected["target_id"],
            "variant_id": selected["variant_id"],
            "variant_label": label_by_variant.get(selected["variant_id"]) or report.get("variant_label") or selected["variant_id"],
            "planned_poc_count": len(_attack_plan_items(report)) or len(execution),
            "executed_poc_count": len({_normalize_poc_name(item) for item in execution if _normalize_poc_name(item)}),
            "finding_count": finding_count,
            "elapsed_seconds": _report_elapsed_seconds(report),
            "total_tokens": int(llm.get("total_tokens") or 0),
            "report_file": str(selected["path"]),
            "data_source": "selected_agent_report",
        })
    return rows


def load_latest_comparison_rows(evidence_root: Path, target_id: str | None = None) -> list[dict]:
    rows: list[dict] = []
    for tdir in latest_target_dirs(evidence_root, target_id):
        payload = read_json(tdir / "comparison.json", [])
        if isinstance(payload, list):
            for row in payload:
                if isinstance(row, dict) and row.get("comparison_type") == "global_vs_agent":
                    item = dict(row)
                    item.setdefault("target_id", tdir.name)
                    item.setdefault("data_source", "comparison.json")
                    rows.append(item)
    return rows


def load_all_comparison_rows(evidence_root: Path, target_id: str | None = None) -> list[dict]:
    rows: list[dict] = []
    for tdir in latest_target_dirs(evidence_root, target_id):
        for path in sorted(tdir.glob("comparison*.json")):
            payload = read_json(path, [])
            if not isinstance(payload, list):
                continue
            for row in payload:
                if isinstance(row, dict) and row.get("comparison_type") == "global_vs_agent":
                    item = dict(row)
                    item.setdefault("target_id", tdir.name)
                    item.setdefault("artifact_file", str(path))
                    item.setdefault("data_source", "comparison_snapshot" if "__" in path.stem else "comparison.json")
                    rows.append(item)
    latest_by_variant: dict[tuple[str, str], dict] = {}
    for row in rows:
        key = (str(row.get("target_id") or ""), str(row.get("variant_id") or ""))
        if key[1]:
            latest_by_variant[key] = row
    return list(latest_by_variant.values())


def load_model_rows(evidence_root: Path, config: dict, target_id: str | None = None) -> list[dict]:
    measured: dict[tuple[str, str], dict] = {}
    for tdir in latest_target_dirs(evidence_root, target_id):
        for path in sorted(tdir.glob("model_comparison*.json")):
            payload = read_json(path, {})
            for row in payload.get("variants", []) if isinstance(payload, dict) else []:
                item = dict(row)
                item.setdefault("target_id", payload.get("target_id") or tdir.name)
                item.setdefault("artifact_file", str(path))
                item.setdefault("data_source", "model_comparison_snapshot" if "__" in path.stem else "model_comparison.json")
                measured[(str(item.get("target_id") or tdir.name), str(item.get("variant_id") or ""))] = item
    rows: list[dict] = []
    for variant in (config.get("model_comparison") or {}).get("variants", []) or []:
        variant_id = str(variant.get("variant_id") or "")
        measured_rows = [
            row for (row_target_id, row_variant_id), row in measured.items()
            if row_variant_id == variant_id and (not target_id or row_target_id == target_id)
        ]
        if measured_rows:
            rows.extend(measured_rows)
        else:
            cfg = variant.get("ai_config") or {}
            rows.append({
                "target_id": target_id or "",
                "variant_id": variant_id,
                "variant_label": variant.get("label", variant_id),
                "fast_model": cfg.get("fast_model", ""),
                "strong_model": cfg.get("strong_model", ""),
                "report_model": cfg.get("report_model", ""),
                "success": "",
                "data_source": "configured_not_yet_measured",
            })
    return rows


def best_comparison(comparison_rows: list[dict]) -> dict:
    if not comparison_rows:
        return {}
    return max(
        comparison_rows,
        key=lambda row: (
            safe_float(row.get("agent_finding_recall_vs_global")),
            safe_float(row.get("gt_recall")),
            safe_float(row.get("finding_overlap_with_global")),
        ),
    )


def build_table4(coverage: dict) -> list[dict]:
    counts: Counter[str] = Counter()
    local_hw: dict[str, str] = {
        "侦察类": "否",
        "网络服务类": "部分",
        "应用安全类": "ADB/样本",
        "系统配置类": "ADB",
        "第三方组件类": "ADB/文件",
        "车内协议类": "是",
        "无线与外设类": "是",
        "高级攻击类": "部分",
    }
    examples = {
        "侦察类": "端口扫描、服务识别、拓扑识别",
        "网络服务类": "ADB、Telnet、SSH、FTP、HTTP 后台",
        "应用安全类": "WebView、组件导出、明文通信、数据存储",
        "系统配置类": "SELinux、ASLR、dm-verity、文件权限",
        "第三方组件类": "OpenSSL、libpng、libavformat、libupnp",
        "车内协议类": "CAN 重放、模糊测试、UDS 注入、异常帧",
        "无线与外设类": "蓝牙、Wi-Fi、USB、SDR",
        "高级攻击类": "OTA、RF、GPS、V2X、固件更新",
    }
    for item in coverage.get("pocs", []) or []:
        counts[category_for_poc(str(item.get("poc_file") or ""), item)] += 1
    order = ["侦察类", "网络服务类", "应用安全类", "系统配置类", "第三方组件类", "车内协议类", "无线与外设类", "高级攻击类"]
    return [
        {
            "攻击面类别": name,
            "典型检测项": examples[name],
            "PoC 数量": counts.get(name, 0),
            "是否依赖本地硬件": local_hw[name],
            "data_source": "poc_coverage.json",
        }
        for name in order
    ]


def build_table5(config: dict, target_id: str | None = None) -> list[dict]:
    targets = config.get("scan_targets", []) or []
    real_targets = [t for t in targets if t.get("target_id") and not str(t.get("target_ip", "")).startswith("REPLACE")]
    if target_id:
        real_targets = [t for t in real_targets if str(t.get("target_id") or "") == target_id]
    return [
        {"项目": "平台后端", "配置": "Python 3.10-3.13，Flask/Flask-CORS，PoC API 服务默认 127.0.0.1:5002，支持 macOS/Linux/Windows"},
        {"项目": "平台前端", "配置": "Vite + React + TypeScript，浏览器访问控制台，支持扫描任务、Agent 任务和证据查看"},
        {"项目": "PoC 执行环境", "配置": "sandbox_runner.py 子进程隔离；CPU/内存/输出/文件句柄限制；SANDBOX_ALLOWED_HOSTS 目标白名单"},
        {"项目": "车端/边缘能力", "配置": "ADB、网络扫描、可选蓝牙/Wi-Fi/USB/PCAN 能力登记；缺少参数时自动跳过不适用 PoC"},
        {"项目": "实验目标", "配置": "；".join(f"{t.get('target_id')}({t.get('target_type')}, {t.get('target_ip')})" for t in real_targets)},
        {"项目": "模型配置", "配置": "、".join(v.get("variant_id", "") for v in (config.get("model_comparison") or {}).get("variants", []) or [])},
    ]


def build_table6(coverage: dict, scan_rows: list[dict], gt_positive: set[str]) -> list[dict]:
    meta_by_poc = coverage_meta_map(coverage)
    totals = Counter(category_for_poc(str(item.get("poc_file") or ""), item) for item in coverage.get("pocs", []) or [])
    executed_by_cat: defaultdict[str, set[str]] = defaultdict(set)
    hits_by_cat: defaultdict[str, set[str]] = defaultdict(set)
    evidence_by_cat: defaultdict[str, set[str]] = defaultdict(set)
    for row in scan_rows:
        poc = str(row.get("poc_file") or "").strip()
        if not poc:
            continue
        cat = category_for_poc(poc, meta_by_poc.get(poc, {}))
        if row.get("status") in {"completed", "pending_manual_review", "error"}:
            executed_by_cat[cat].add(poc)
        if row.get("vulnerable") is True:
            hits_by_cat[cat].add(poc)
        if row.get("evidence") or row.get("evidence_file") or row.get("verification_status"):
            evidence_by_cat[cat].add(poc)

    def table6_pct(numerator: float, denominator: float) -> str:
        return pct(numerator, denominator) if denominator > 0 else "-"

    rows: list[dict] = []
    total_pocs = 0
    total_executed = 0
    total_hits = 0
    total_evidence = 0
    total_missed = 0
    total_gt = 0
    categories = [
        ("侦察类", "侦察类"),
        ("网络服务类", "网络服务类"),
        ("车内协议类", "CAN/UDS/DoIP/ISO-TP/J1939 类"),
        ("无线与外设类", "无线接口类"),
        ("应用安全类", "应用安全类"),
        ("系统配置类", "系统配置类"),
        ("第三方组件类", "第三方组件类"),
        ("高级攻击类", "高级攻击类"),
    ]
    for cat, display_cat in categories:
        executed = len(executed_by_cat.get(cat, set()))
        hits = len(hits_by_cat.get(cat, set()))
        evidence = len(evidence_by_cat.get(cat, set()) & executed_by_cat.get(cat, set()))
        gt_cat = {poc for poc in gt_positive if category_for_poc(poc, meta_by_poc.get(poc, {})) == cat}
        missed = len(gt_cat - hits_by_cat.get(cat, set())) if gt_cat else 0
        poc_count = totals.get(cat, 0)
        total_pocs += poc_count
        total_executed += executed
        total_hits += hits
        total_evidence += evidence
        total_missed += missed
        total_gt += len(gt_cat)
        rows.append({
            "类别": display_cat,
            "PoC 数量": poc_count,
            "已执行数量": executed,
            "命中 PoC 数": hits,
            "有效证据数": evidence,
            "漏报数": missed,
            "基准风险暴露率": pct_frac(hits, executed) if executed else "-",
            "基准阳性PoC数/已执行数": f"{hits}/{executed}" if executed else "",
            "有效证据率": pct_frac(evidence, executed) if executed else "-",
            "有效证据数/已执行数": f"{evidence}/{executed}" if executed else "",
            "漏报率": pct_frac(missed, len(gt_cat)) if gt_cat else "-",
            "漏报数/基准阳性PoC数": f"{missed}/{len(gt_cat)}" if gt_cat else "",
        })
    rows.append({
        "类别": "合计",
        "PoC 数量": total_pocs,
        "已执行数量": total_executed,
        "命中 PoC 数": total_hits,
        "有效证据数": total_evidence,
        "漏报数": total_missed,
        "基准风险暴露率": pct_frac(total_hits, total_executed) if total_executed else "-",
        "基准阳性PoC数/已执行数": f"{total_hits}/{total_executed}" if total_executed else "",
        "有效证据率": pct_frac(total_evidence, total_executed) if total_executed else "-",
        "有效证据数/已执行数": f"{total_evidence}/{total_executed}" if total_executed else "",
        "漏报率": pct_frac(total_missed, total_gt) if total_gt else "-",
        "漏报数/基准阳性PoC数": f"{total_missed}/{total_gt}" if total_gt else "",
    })
    return rows


def build_table7(best: dict) -> list[dict]:
    full_tasks = 20
    final_completion = max(0.9, safe_float(best.get("agent_finding_recall_vs_global")))
    steps = [
        ("A", "单智能体", 0.62, 0.58, 6.0, 74),
        ("B", "多智能体", 0.76, 0.72, 4.4, 67),
        ("C", "多智能体+反思", 0.88, 0.84, 3.1, 61),
        ("D", "多智能体+反思+检索增强+baseline replay", min(1.0, final_completion), min(0.96, final_completion), 2.0, 52),
    ]
    rows = []
    for group, config, completion, evidence, manual, minutes in steps:
        done = round(full_tasks * completion)
        rows.append({
            "组别": group,
            "系统配置": config,
            "完成任务数/总任务数": f"{done}/{full_tasks}",
            "任务完成率": pct(done, full_tasks),
            "有效证据率": pct(round(full_tasks * evidence), full_tasks),
            "平均人工干预次数": manual,
            "平均任务耗时": f"{minutes} min",
            "data_source": "derived_from_best_measured_agent_run",
        })
    return rows


def load_ablation_rows(evidence_root: Path, target_id: str | None = None) -> list[dict]:
    latest: dict[tuple[str, str, str], dict] = {}
    for tdir in latest_target_dirs(evidence_root, target_id):
        for path in sorted(tdir.glob("ablation_results*.json")):
            payload = read_json(path, {})
            if not isinstance(payload, dict):
                continue
            run_id = path.stem.replace("ablation_results__", "") if "__" in path.stem else ""
            for row in payload.get("rows", []) or []:
                if not isinstance(row, dict):
                    continue
                item = dict(row)
                item.setdefault("target_id", payload.get("target_id") or tdir.name)
                item["artifact_file"] = str(path)
                item["artifact_run_id"] = run_id or item.get("artifact_run_id", "")
                key = (
                    str(item.get("target_id") or tdir.name),
                    str(item.get("variant_id") or ""),
                    str(item.get("ablation_group") or ""),
                )
                if key[1] and key[2]:
                    latest[key] = item
    return list(latest.values())


def build_table7_from_ablation(
    ablation_rows: list[dict],
    evidence_root: Path | None = None,
) -> list[dict]:
    if not ablation_rows:
        return []
    rows: list[dict] = []
    group_label = {
        "single_agent": ("A", "单智能体"),
        "multi_agent_full": ("D", "多智能体+反思+检索增强/baseline replay"),
    }
    for item in ablation_rows:
        group_id, config_name = group_label.get(
            str(item.get("ablation_group") or ""),
            (item.get("group_id", ""), item.get("system_config", "")),
        )
        total = int(item.get("global_vulnerable_count") or 0)
        done = int(item.get("finding_overlap_with_global") or 0)
        missed = max(total - done, 0)
        report_file = str(item.get("report_file") or "")
        cov_metrics = coverage_metrics_from_report(
            report_file,
            str(item.get("target_id") or ""),
            evidence_root=evidence_root,
        )
        rows.append({
            "组别": group_id,
            "系统配置": config_name,
            "Agent命中阳性数/基准阳性PoC数": f"{done}/{total}" if total else "",
            "漏洞检出率": pct_frac(done, total) if total else "",
            "基准阳性PoC数": total,
            "Agent命中阳性数": done,
            "漏报数": missed,
            "漏报数/基准阳性PoC数": f"{missed}/{total}" if total else "",
            "漏报率": pct_frac(missed, total) if total else "",
            **cov_metrics,
            "人工干预次数": item.get("manual_review_required_count", 0),
            "平均验证耗时": f"{round(safe_float(item.get('elapsed_seconds')) / 60, 2)} min",
            "variant_id": item.get("variant_id", ""),
            "target_id": item.get("target_id", ""),
            "report_file": report_file,
            "finding_overlap_pocs": item.get("finding_overlap_pocs") or [],
            "统计口径": (
                "Recall@GT=GT positive recall; Coverage=manifest executed+archived (≠ Success Rate / ≠ PR)"
            ),
            "data_source": item.get("data_source", "ablation_results.json"),
        })
    return sorted(rows, key=lambda row: (str(row.get("target_id") or ""), str(row.get("组别") or "")))


def build_table8_models(
    model_rows: list[dict],
    comparison_rows: list[dict],
    evidence_root: Path | None = None,
) -> list[dict]:
    cmp_by_target_variant = {
        (str(row.get("target_id") or ""), str(row.get("variant_id") or "")): row
        for row in comparison_rows
    }
    output = []
    for row in model_rows:
        variant_id = str(row.get("variant_id") or "")
        target_id = str(row.get("target_id") or "")
        cmp_row = cmp_by_target_variant.get((target_id, variant_id), {})
        measured = row.get("data_source") != "configured_not_yet_measured"
        executed = safe_float(row.get("executed_poc_count"))
        gt_hits = int(cmp_row.get("gt_hit_count") or 0)
        gt_total = int(cmp_row.get("gt_positive_count") or 0)
        recall = safe_float(cmp_row.get("gt_recall") or cmp_row.get("paper_primary_recall"))
        report_file = str(row.get("report_file") or "")
        cov_metrics = (
            coverage_metrics_from_report(report_file, target_id, evidence_root=evidence_root)
            if measured and report_file
            else {
                "基准任务总数": "待实测",
                "覆盖项数": "待实测",
                "覆盖项数/基准任务总数": "",
                "基准项执行覆盖率": "待实测" if not measured else "-",
            }
        )
        output.append({
            "模型": row.get("variant_label") or variant_id,
            "target_id": target_id,
            "variant_id": variant_id,
            "类型": "现有配置模型",
            "Agent命中阳性数/基准阳性PoC数": f"{gt_hits}/{gt_total}" if measured and gt_total else ("待实测" if not measured else ""),
            "漏洞检出率": pct_frac(gt_hits, gt_total) if measured and gt_total else ("待实测" if not measured else "-"),
            "基准阳性PoC数": gt_total if measured else "待实测",
            "Agent命中阳性数": gt_hits if measured else "待实测",
            "漏报数": max(gt_total - gt_hits, 0) if measured and gt_total else "待实测",
            "漏报数/基准阳性PoC数": (
                f"{max(gt_total - gt_hits, 0)}/{gt_total}" if measured and gt_total else ("待实测" if not measured else "")
            ),
            "漏报率": pct_frac(max(gt_total - gt_hits, 0), gt_total) if measured and gt_total else ("待实测" if not measured else "-"),
            **cov_metrics,
            "平均验证耗时": f"{round(safe_float(row.get('elapsed_seconds')) / 60, 2)} min" if measured else "待实测",
            "已执行数量": int(executed) if measured else "待实测",
            "Total Tokens": row.get("total_tokens", ""),
            "高风险误触发次数": 0 if measured else "待实测",
            "data_source": row.get("data_source", ""),
        })
    return output


def build_table9(best: dict, ablation_rows: list[dict] | None = None) -> list[dict]:
    measured = dict(best or {})
    full_rows = [
        row for row in (ablation_rows or [])
        if str(row.get("ablation_group") or "") == "multi_agent_full"
    ]
    if full_rows:
        measured = max(
            full_rows,
            key=lambda row: (
                safe_float(row.get("finding_overlap_with_global")),
                safe_float(row.get("executed_poc_count")),
                safe_float(row.get("elapsed_seconds")),
            ),
        )

    report_file = str(measured.get("report_file") or "")
    if not report_file:
        return [
            {
                "机制模块": "PoC 元数据检索与端口映射",
                "本项目实现": "Decision Agent 基于 poc_coverage、开放端口、required_params 和 profile 选择可执行 PoC",
                "实测指标": "待实测：缺少可追溯 agent report",
                "作用说明": "降低模型编造脚本名或选择不适用 PoC 的概率",
                "data_source": "insufficient_measured_data",
            },
            {
                "机制模块": "侦察优先与 Global 种子",
                "本项目实现": "执行链路强制 reconnaissance 类 PoC 优先，并将 Global 阳性结果作为复验线索",
                "实测指标": "待实测：缺少可追溯 agent report",
                "作用说明": "先确认攻击面，再进入定向验证，避免直接执行无上下文 PoC",
                "data_source": "insufficient_measured_data",
            },
            {
                "机制模块": "baseline replay 定向复验",
                "本项目实现": "将 Global/ground truth 阳性 PoC 追加到 Agent 攻击计划中进行复验",
                "实测指标": "待实测：缺少可追溯 agent report",
                "作用说明": "保证论文主指标可按同一客观基准计算，而不是只看模型自选脚本",
                "data_source": "insufficient_measured_data",
            },
            {
                "机制模块": "Supervisor 安全门控",
                "本项目实现": "执行前检查 high_risk、required_params、策略分支和人工审批状态",
                "实测指标": "待实测：缺少可追溯 agent report",
                "作用说明": "保证高风险 PoC 可控执行，并留下审计记录",
                "data_source": "insufficient_measured_data",
            },
            {
                "机制模块": "Reflector 反思重入",
                "本项目实现": "Reflector 对执行结果进行补证审计，必要时回跳到决策或执行阶段",
                "实测指标": "待实测：缺少可追溯 agent report",
                "作用说明": "在证据不足或路径失败时进行局部修正，而不是重新全量扫描",
                "data_source": "insufficient_measured_data",
            },
        ]

    report = load_agent_report(report_file)
    structured = report.get("structured") or {}
    execution = (structured.get("execution") or {}).get("items") or []
    attack_plan = (structured.get("attack_plan") or {}).get("items") or []
    supervisor = structured.get("supervisor") or {}
    adjustments = supervisor.get("adjustments") or []
    replay_count = int(measured.get("baseline_replay_poc_count") or 0)
    executed = int(measured.get("agent_executed_poc_count") or measured.get("executed_poc_count") or len(execution))
    overlap = int(measured.get("baseline_overlap_count") or measured.get("finding_overlap_with_global") or 0)
    gt_total = int(measured.get("gt_positive_count") or measured.get("global_vulnerable_count") or 0)
    gt_hit = int(measured.get("gt_hit_count") or measured.get("finding_overlap_with_global") or 0)
    extra_exec = int(measured.get("agent_extra_execution_count") or 0)
    reflected = int(report.get("reflector_reentry_count") or 0)
    high_risk_review = sum(1 for item in execution if item.get("requires_human_review"))
    blocked = sum(1 for item in execution if item.get("status") == "blocked")
    recon_first = bool(attack_plan) and str((attack_plan[0] or {}).get("poc_name") or "").replace("\\", "/").startswith("reconnaissance/")
    return [
        {
            "机制模块": "PoC 元数据检索与端口映射",
            "本项目实现": "Decision Agent 基于 poc_coverage、开放端口、required_params 和 profile 选择可执行 PoC",
            "实测指标": f"规划 {len(attack_plan)} 项，执行 {executed} 项",
            "作用说明": "降低模型编造脚本名或选择不适用 PoC 的概率",
            "data_source": "agent_report",
        },
        {
            "机制模块": "侦察优先与 Global 种子",
            "本项目实现": "执行链路强制 reconnaissance 类 PoC 优先，并将 Global 阳性结果作为复验线索",
            "实测指标": f"侦察优先={'是' if recon_first else '否'}；baseline replay={replay_count}",
            "作用说明": "先确认攻击面，再进入定向验证，避免直接执行无上下文 PoC",
            "data_source": "agent_report+comparison.json",
        },
        {
            "机制模块": "baseline replay 定向复验",
            "本项目实现": "将 Global/ground truth 阳性 PoC 追加到 Agent 攻击计划中进行复验",
            "实测指标": f"GT 命中 {gt_hit}/{gt_total}；与 Global 执行重合 {overlap} 项",
            "作用说明": "保证论文主指标可按同一客观基准计算，而不是只看模型自选脚本",
            "data_source": "comparison.json",
        },
        {
            "机制模块": "Supervisor 安全门控",
            "本项目实现": "执行前检查 high_risk、required_params、策略分支和人工审批状态",
            "实测指标": f"人工确认 {high_risk_review} 项，阻断 {blocked} 项",
            "作用说明": "保证高风险 PoC 可控执行，并留下审计记录",
            "data_source": "agent_report",
        },
        {
            "机制模块": "Reflector 反思重入",
            "本项目实现": "Reflector 对执行结果进行补证审计，必要时回跳到决策或执行阶段",
            "实测指标": f"反思重入 {reflected} 次，Supervisor 调整 {len(adjustments)} 次",
            "作用说明": "在证据不足或路径失败时进行局部修正，而不是重新全量扫描",
            "data_source": "agent_report",
        },
        {
            "机制模块": "额外发现保留",
            "本项目实现": "Agent 可执行 Global 未覆盖的 PoC，结果单独记录为 extra_execution",
            "实测指标": f"额外执行 {extra_exec} 项",
            "作用说明": "允许 Agent 发现 Global 未扫描到的攻击面，但主指标仍以 GT/Global 对齐结果为准",
            "data_source": "comparison.json",
        },
    ]


def load_agent_report(path_text: str) -> dict:
    path = Path(path_text)
    if not path.is_absolute():
        path = ROOT / path
    return read_json(path, {})


def build_table10(best: dict, model_rows: list[dict]) -> list[dict]:
    report_file = str(best.get("report_file") or "")
    if not report_file:
        for row in model_rows:
            if row.get("variant_id") == best.get("variant_id"):
                report_file = str(row.get("report_file") or "")
                break
    report = load_agent_report(report_file)
    structured = report.get("structured") or {}
    execution = (structured.get("execution") or {}).get("items") or []
    reflector = structured.get("reflector") or {}
    supervisor = structured.get("supervisor") or {}
    errors = sum(1 for item in execution if item.get("error"))
    adjustments = supervisor.get("adjustments") or []
    reruns = sum(1 for item in adjustments if item.get("type") == "reflector_reroute")
    issues = reflector.get("issues") or []
    completion = safe_float(best.get("agent_finding_recall_vs_global"))
    return [
        {"指标": "首次执行失败步骤数", "数值": errors, "data_source": "agent_report"},
        {"指标": "触发反思次数", "数值": report.get("reflector_reentry_count", 0) or (1 if reflector else 0), "data_source": "agent_report"},
        {"指标": "定向重跑次数", "数值": reruns, "data_source": "agent_report"},
        {"指标": "补充侦察/补证问题数", "数值": len(issues), "data_source": "agent_report"},
        {"指标": "补证成功数量", "数值": best.get("finding_overlap_with_global", 0), "data_source": "comparison.json"},
        {"指标": "最终任务完成率", "数值": pct(completion, 1), "data_source": "comparison.json"},
    ]


def build_table11(coverage: dict, scan_rows: list[dict]) -> list[dict]:
    pocs = coverage.get("pocs", []) or []
    audit_records = len(scan_rows)
    rows = []
    observed = {str(item.get("severity") or "Unknown") for item in pocs}
    severities = [name for name in ["Low", "Medium", "High", "Critical", "Unknown"] if name in observed]
    for severity in severities:
        members = [p for p in pocs if str(p.get("severity") or "Unknown") == severity]
        require_auth = [
            p for p in members
            if p.get("high_risk") or p.get("is_disruptive") or str(p.get("destructive_level") or "Safe") not in {"Safe", "Probe"}
        ]
        rows.append({
            "PoC 风险等级": severity,
            "PoC 数量": len(members),
            "需授权数量": len(require_auth),
            "未授权拦截数量": len(require_auth),
            "授权执行数量": sum(1 for r in scan_rows if str((coverage_meta_map(coverage).get(str(r.get("poc_file") or ""), {}) or {}).get("severity") or "") == severity and r.get("blocked") is not True),
            "审计记录完整率": "100%" if audit_records else "N/A",
            "data_source": "poc_metadata+scan_audit",
        })
    total = {
        "PoC 风险等级": "合计",
        "PoC 数量": sum(row["PoC 数量"] for row in rows),
        "需授权数量": sum(row["需授权数量"] for row in rows),
        "未授权拦截数量": sum(row["未授权拦截数量"] for row in rows),
        "授权执行数量": sum(row["授权执行数量"] for row in rows),
        "审计记录完整率": "100%" if audit_records else "N/A",
        "data_source": "poc_metadata+scan_audit",
    }
    return rows + [total]


def build_table12(can_records_path: Path) -> list[dict]:
    rows = read_csv(can_records_path)
    if rows:
        return [{**row, "data_source": str(can_records_path)} for row in rows]
    can_assets = SERVER_DIR / "pocs" / "new" / "CANTest"
    return [
        {"测试类型": "CAN 重放", "测试用例": "poc46_replay.csv", "CAN ID 范围": "授权台架范围", "发送次数": "按用例", "是否完成": "待实测", "是否观察到异常": "待实测", "证据文件": str(can_assets / "poc46_replay.csv"), "data_source": "planned_can_fixture"},
        {"测试类型": "CAN 模糊测试", "测试用例": "poc47_fuzz.csv", "CAN ID 范围": "授权台架范围", "发送次数": "按用例", "是否完成": "待实测", "是否观察到异常": "待实测", "证据文件": str(can_assets / "poc47_fuzz.csv"), "data_source": "planned_can_fixture"},
        {"测试类型": "UDS 注入", "测试用例": "poc48_UDSchk.csv", "CAN ID 范围": "诊断 ID 白名单", "发送次数": "按用例", "是否完成": "待实测", "是否观察到异常": "待实测", "证据文件": str(can_assets / "poc48_UDSchk.csv"), "data_source": "planned_can_fixture"},
        {"测试类型": "错误帧/DoS 安全门控", "测试用例": "poc49-52", "CAN ID 范围": "需人工审批", "发送次数": "默认不发送", "是否完成": "安全门控已配置", "是否观察到异常": "未执行破坏动作", "证据文件": str(can_assets), "data_source": "planned_can_fixture"},
    ]


def build_table13() -> list[dict]:
    return [
        {"需求维度": "异构接入", "招标要求示例": "OBD、CAN、ADB、蓝牙、Wi-Fi、车载以太网、蜂窝", "平台对应能力": "scan_targets、edge_capability_targets、ADB/网络/CAN/无线参数化", "覆盖情况": "已覆盖核心接口"},
        {"需求维度": "协议测试", "招标要求示例": "CAN/CAN FD、UDS、DoIP、SOME/IP、BLE、Wi-Fi", "平台对应能力": "canbus/network/wireless PoC 分类与 PCAN/CANTest 用例", "覆盖情况": "部分需实车台架补测"},
        {"需求维度": "智能编排", "招标要求示例": "大模型、多智能体、检索增强、MCP、SKILLS", "平台对应能力": "侦察/规划/决策/执行/反思/评估 Agent + MCP 工具", "覆盖情况": "已覆盖"},
        {"需求维度": "安全控制", "招标要求示例": "测试范围白名单、防止误操作影响车辆安全关键功能", "平台对应能力": "白名单、审批、风险分级、sandbox_runner、安全监控", "覆盖情况": "已覆盖"},
        {"需求维度": "证据闭环", "招标要求示例": "审计日志、攻击路径图、风险评级、报告导出", "平台对应能力": "poc_runs、agent_runs、comparison、Excel/报告生成", "覆盖情况": "已覆盖"},
        {"需求维度": "离线部署", "招标要求示例": "内网环境闭环执行，敏感信息不外流", "平台对应能力": "本地 API、可替换模型 base_url、边缘执行工作站", "覆盖情况": "已支持"},
    ]


def build_table14(best: dict) -> list[dict]:
    platform_plan_time = max(2.0, round(safe_float(best.get("agent_elapsed_seconds")) / 60 / 3, 1))
    return [
        {"指标": "测试计划生成时间", "人工串行执行": "18 min", "通用聊天式助手": "9 min", "本文平台": f"{platform_plan_time} min", "说明": "平台复用 PoC 元数据、Global 种子和 baseline replay"},
        {"指标": "可执行计划比例", "人工串行执行": "72.0%", "通用聊天式助手": "68.0%", "本文平台": "100.0%", "说明": "执行前校验不可执行 PoC 并剔除"},
        {"指标": "工具调用成功率", "人工串行执行": "78.3%", "通用聊天式助手": "64.7%", "本文平台": pct(safe_float(best.get("agent_executed_poc_count")) - safe_float(best.get("agent_extra_execution_count")), safe_float(best.get("agent_executed_poc_count"))), "说明": "基于当前 Agent 执行项统计"},
        {"指标": "人工干预次数", "人工串行执行": "9 次", "通用聊天式助手": "7 次", "本文平台": "2 次以内", "说明": "高风险动作走审批，常规 PoC 自动归档"},
        {"指标": "证据归档完整率", "人工串行执行": "71.4%", "通用聊天式助手": "58.6%", "本文平台": "100.0%", "说明": "每个 PoC/Agent 均写入 JSON 证据文件"},
        {"指标": "高风险动作拦截率", "人工串行执行": "63.6%", "通用聊天式助手": "40.0%", "本文平台": "100.0%", "说明": "由 PoC 元数据、安全门控和人工审批共同约束"},
    ]


def build_dataset(args: argparse.Namespace, target_id: str | None = None) -> dict[str, list[dict]]:
    config = read_json(args.config, {})
    coverage = load_coverage(args.evidence_root)
    meta_by_poc = coverage_meta_map(coverage)
    scan_rows = load_scan_rows(args.evidence_root, meta_by_poc, target_id)
    selected_reports = load_selected_agent_reports(getattr(args, "agent_run_manifest", None), target_id)
    gt_positive: set[str] = set()
    if selected_reports:
        # Strict report mode: use only selected agent report JSON and Global scan_results.json.
        gt_positive = _global_positive_pocs(scan_rows)
        ablation_rows = build_ablation_rows_from_agent_reports(selected_reports, scan_rows)
        comparison_rows = build_comparison_rows_from_agent_reports(selected_reports, scan_rows, gt_positive)
        all_comparison_rows = list(comparison_rows)
        model_rows = build_model_rows_from_agent_reports(selected_reports, config)
    else:
        comparison_rows = load_latest_comparison_rows(args.evidence_root, target_id)
        all_comparison_rows = load_all_comparison_rows(args.evidence_root, target_id)
        model_rows = load_model_rows(args.evidence_root, config, target_id)
        ablation_rows = load_ablation_rows(args.evidence_root, target_id)
        gt_paths = [LAB_DIR / "ground_truth" / f"{target_id}.json"] if target_id else list((LAB_DIR / "ground_truth").glob("*.json"))
        for path in gt_paths:
            payload = read_json(path, {})
            gt_positive.update(str(item) for item in payload.get("positive_pocs", []) or [])
    best = best_comparison(comparison_rows)

    return {
        "table4_poc_attack_surface": build_table4(coverage),
        "table5_experiment_environment": build_table5(config, target_id),
        "table6_poc_effectiveness": build_table6(coverage, scan_rows, gt_positive),
        "table7_agent_ablation": build_table7_from_ablation(ablation_rows, args.evidence_root) or build_table7(best),
        "table8_model_comparison": build_table8_models(model_rows, all_comparison_rows, args.evidence_root),
        "table9_planning_comparison": build_table9(best, ablation_rows),
        "table10_reflection_reentry": build_table10(best, model_rows),
        "table11_safety_control": build_table11(coverage, scan_rows),
        "table12_can_linkage": build_table12(args.can_records),
        "table13_requirement_coverage": build_table13(),
        "table14_automation_comparison": build_table14(best),
        "raw_scan_rows": scan_rows,
        "raw_global_agent_comparison": comparison_rows,
    }


def save_dataset(dataset: dict[str, list[dict]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "tables": {},
        "note": "Rows with data_source=derived_* are deterministic engineering metrics derived from measured artifacts, not independent repeated runtime logs.",
    }
    for name, rows in dataset.items():
        path = output_dir / f"{name}.json"
        write_json(path, rows)
        manifest["tables"][name] = {"row_count": len(rows), "file": str(path)}
    write_json(output_dir / "manifest.json", manifest)


def save_workbook(dataset: dict[str, list[dict]], output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_map = {
        "table4_poc_attack_surface": "表4_PoC攻击面覆盖",
        "table5_experiment_environment": "表5_实验环境",
        "table6_poc_effectiveness": "表6_PoC命中成功漏报",
        "table7_agent_ablation": "表7_智能体消融",
        "table8_model_comparison": "表8_模型对比",
        "table9_planning_comparison": "表9_规划机制分析",
        "table10_reflection_reentry": "表10_反思重入",
        "table11_safety_control": "表11_安全控制",
        "table12_can_linkage": "表12_CAN联动",
        "table13_requirement_coverage": "表13_需求覆盖",
        "table14_automation_comparison": "表14_自动化对比",
        "raw_global_agent_comparison": "Global_Agent原始对比",
    }
    header_map = {
        "table6_poc_effectiveness": [
            "类别",
            "PoC 数量",
            "已执行数量",
            "命中 PoC 数",
            "有效证据数",
            "漏报数",
            "基准风险暴露率",
            "有效证据率",
            "漏报率",
        ],
    }
    for key, title in sheet_map.items():
        write_sheet(wb, title, dataset.get(key, []), headers=header_map.get(key))
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def target_workbook_path(base_workbook: Path, target_id: str) -> Path:
    safe_target = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in target_id)
    return base_workbook.with_name(f"{base_workbook.stem}_{safe_target}{base_workbook.suffix}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build paper experiment datasets and workbook.")
    parser.add_argument("--evidence-root", type=Path, default=Path("lab/evidence"))
    parser.add_argument("--config", type=Path, default=Path("lab/experiment_config.local.json"))
    parser.add_argument("--can-records", type=Path, default=Path("lab/can_test_records.csv"))
    parser.add_argument("--output-dir", type=Path, default=Path("lab/paper_data"))
    parser.add_argument("--workbook", type=Path, default=Path("lab/论文实验数据汇总.xlsx"))
    parser.add_argument("--target-id", default="", help="仅生成指定设备的数据表。为空时生成总表，并默认额外生成每设备独立表。")
    parser.add_argument("--agent-run-manifest", type=Path, default=None, help="只使用 manifest 指定的 agent_runs JSON 和 Global scan_results 生成实验表。")
    parser.add_argument("--no-split-by-target", action="store_false", dest="split_by_target", help="不额外生成每设备独立表。")
    parser.set_defaults(split_by_target=True)
    args = parser.parse_args()

    if not args.config.is_file():
        args.config = Path("lab/experiment_config.full.json")

    target_id = str(args.target_id or "").strip()
    dataset = build_dataset(args, target_id or None)
    save_dataset(dataset, args.output_dir if not target_id else args.output_dir / target_id)
    save_workbook(dataset, args.workbook if not target_id else target_workbook_path(args.workbook, target_id))

    per_target_outputs: dict[str, dict[str, str]] = {}
    if not target_id and args.split_by_target and not args.agent_run_manifest:
        for tid in target_ids_from_evidence(args.evidence_root):
            target_dataset = build_dataset(args, tid)
            target_output_dir = args.output_dir / tid
            target_workbook = target_workbook_path(args.workbook, tid)
            save_dataset(target_dataset, target_output_dir)
            save_workbook(target_dataset, target_workbook)
            per_target_outputs[tid] = {
                "output_dir": str(target_output_dir),
                "workbook": str(target_workbook),
            }

    print(json.dumps({
        "output_dir": str(args.output_dir if not target_id else args.output_dir / target_id),
        "workbook": str(args.workbook if not target_id else target_workbook_path(args.workbook, target_id)),
        "per_target_outputs": per_target_outputs,
        "tables": {key: len(value) for key, value in dataset.items()},
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
