#!/usr/bin/env python3
"""合并各 target 证据，生成论文用总表 lab/论文实验数据汇总.xlsx"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

LAB_DIR = Path(__file__).resolve().parent
if str(LAB_DIR) not in sys.path:
    sys.path.insert(0, str(LAB_DIR))

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from artifacts import load_latest_json_artifact, load_versioned_json_rows


def sanitize_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    return text[:32767] if len(text) > 32767 else text


def read_json(path: Path, default: Any):
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]):
    ws = wb.create_sheet(title)
    ws.append([sanitize_excel_value(header) for header in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([sanitize_excel_value(row.get(header, "")) for header in headers])
    for index, header in enumerate(headers, start=1):
        width = max(12, len(str(header)) + 2)
        for row_index in range(2, min(ws.max_row, 120) + 1):
            width = max(width, min(45, len(str(ws.cell(row_index, index).value or "")) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def discover_targets(evidence_root: Path) -> list[str]:
    targets = []
    for path in sorted(evidence_root.iterdir()):
        if path.is_dir() and (
            (path / "scan_results.json").is_file()
            or any(path.glob("scan_results__*.json"))
        ):
            targets.append(path.name)
    return targets


def main() -> int:
    parser = argparse.ArgumentParser(description="Build merged paper workbook from all target evidence dirs.")
    parser.add_argument("--evidence-root", type=Path, default=Path("lab/evidence"))
    parser.add_argument("--config", type=Path, default=Path("lab/experiment_config.local.json"))
    parser.add_argument("--output", type=Path, default=Path("lab/论文实验数据汇总.xlsx"))
    parser.add_argument(
        "--use-cached-coverage",
        action="store_true",
        help="使用 evidence 目录下已缓存的 poc_coverage.json（默认每次从 PoC 源码重新扫描）",
    )
    args = parser.parse_args()

    if not args.config.is_file():
        args.config = Path("lab/experiment_config.full.json")
    config = read_json(args.config, {})
    typical_cases = config.get("typical_cases") or []

    if getattr(args, "use_cached_coverage", False):
        coverage = load_latest_json_artifact(
            args.evidence_root, "poc_coverage.json",
            {"by_category": {}, "by_attack_surface": {}, "total": 0},
        )
    else:
        from run_experiment import collect_poc_coverage, write_json

        coverage = collect_poc_coverage()
        write_json(args.evidence_root / "poc_coverage.json", coverage)

    all_scan: list[dict] = []
    all_agent: list[dict] = []
    all_model: list[dict] = []
    all_compare: list[dict] = []
    all_can: list[dict] = []
    all_blocked: list[dict] = []

    for target_id in discover_targets(args.evidence_root):
        tdir = args.evidence_root / target_id
        for row in load_latest_json_artifact(tdir, "scan_results.json", []):
            item = dict(row)
            item["目标"] = target_id
            all_scan.append(item)
            if item.get("blocked") or item.get("requires_approval"):
                all_blocked.append({
                    "目标": target_id,
                    "PoC编号": item.get("poc_display_id") or item.get("poc_file"),
                    "执行状态": item.get("status"),
                    "是否拦截": item.get("blocked"),
                    "证据文件": item.get("evidence_file"),
                })

        for row in load_versioned_json_rows(tdir, "agent_orchestration.json"):
            item = dict(row)
            item["目标"] = target_id
            all_agent.append(item)

        for row in load_versioned_json_rows(tdir, "model_comparison.json", payload_key="variants"):
            item = dict(row)
            item["目标"] = target_id
            all_model.append(item)

        compare_rows: list[dict] = []
        latest_cmp = read_json(tdir / "comparison.json", [])
        if isinstance(latest_cmp, list):
            compare_rows = [
                dict(row) for row in latest_cmp
                if isinstance(row, dict) and row.get("comparison_type") == "global_vs_agent"
            ]
        if not compare_rows:
            latest_snapshot = load_latest_json_artifact(tdir, "comparison.json", [])
            if isinstance(latest_snapshot, list):
                compare_rows = [
                    dict(row) for row in latest_snapshot
                    if isinstance(row, dict) and row.get("comparison_type") == "global_vs_agent"
                ]
        if not compare_rows and (
            (tdir / "model_comparison.json").is_file()
            or any(tdir.glob("model_comparison__*.json"))
        ):
            try:
                from compare_global_vs_agent import build_comparison

                generated = build_comparison(tdir, target_id)
                compare_rows = [
                    dict(row) for row in generated
                    if row.get("comparison_type") == "global_vs_agent"
                ]
            except Exception:
                pass
        all_compare.extend(compare_rows)

        can_path = tdir / "can_test_records.csv"
        for row in read_csv(can_path):
            item = dict(row)
            item["目标"] = target_id
            all_can.append(item)

    wb = Workbook()
    wb.remove(wb.active)

    cov_rows = []
    for name, count in sorted((coverage.get("by_category") or {}).items()):
        cov_rows.append({"统计维度": "PoC分类", "名称": name, "数量": count})
    for name, count in sorted((coverage.get("by_attack_surface") or {}).items()):
        cov_rows.append({"统计维度": "攻击面", "名称": name, "数量": count})
    cov_rows.append({"统计维度": "总计", "名称": "PoC总数", "数量": coverage.get("total", 0)})
    write_sheet(wb, "表1_PoC覆盖情况", ["统计维度", "名称", "数量"], cov_rows)

    write_sheet(wb, "表2_扫描执行结果", [
        "目标", "目标类型", "PoC编号", "检测项", "类别", "攻击面", "执行状态", "耗时s", "是否发现风险", "人工确认", "确认状态", "人工等待s", "是否拦截", "证据文件",
    ], [{
        "目标": r.get("目标"),
        "目标类型": r.get("target_type"),
        "PoC编号": r.get("poc_display_id") or r.get("poc_file"),
        "检测项": r.get("poc_name"),
        "类别": r.get("category"),
        "攻击面": r.get("attack_surface"),
        "执行状态": r.get("status"),
        "耗时s": r.get("elapsed_seconds"),
        "是否发现风险": r.get("vulnerable"),
        "人工确认": r.get("requires_human_review"),
        "确认状态": r.get("verification_status"),
        "人工等待s": r.get("manual_review_wait_seconds"),
        "是否拦截": r.get("blocked"),
        "证据文件": r.get("evidence_file"),
    } for r in all_scan])

    write_sheet(wb, "表3_多Agent编排", [
        "目标", "目标类型", "artifact_run_id", "任务ID", "模型变体", "fast_model", "strong_model", "规划PoC", "执行PoC", "发现数", "待人工确认数", "需人工确认数", "人工等待s", "LLM调用数", "Total Tokens", "相对Global执行覆盖率", "反思重入", "耗时s",
    ], [{
        "目标": r.get("目标"),
        "目标类型": r.get("target_type"),
        "artifact_run_id": r.get("artifact_run_id", ""),
        "任务ID": r.get("task_id"),
        "模型变体": r.get("variant_label") or r.get("variant_id"),
        "fast_model": r.get("fast_model"),
        "strong_model": r.get("strong_model"),
        "规划PoC": r.get("planned_poc_count"),
        "执行PoC": r.get("executed_poc_count"),
        "发现数": r.get("finding_count"),
        "待人工确认数": r.get("manual_review_pending_count"),
        "需人工确认数": r.get("manual_review_required_count"),
        "人工等待s": r.get("manual_review_wait_seconds"),
        "LLM调用数": r.get("llm_call_count"),
        "Total Tokens": r.get("total_tokens"),
        "相对Global执行覆盖率": r.get("agent_execution_coverage_vs_global"),
        "反思重入": r.get("reflection_reentry_count"),
        "耗时s": r.get("elapsed_seconds"),
    } for r in all_agent])

    write_sheet(wb, "表4_CAN网关联动", [
        "目标", "case_id", "test_type", "interface", "can_id", "send_count", "observed_response", "blocked_by_safety", "evidence_file",
    ], [{
        "目标": r.get("目标"),
        "case_id": r.get("case_id"),
        "test_type": r.get("test_type"),
        "interface": r.get("interface"),
        "can_id": r.get("can_id"),
        "send_count": r.get("send_count"),
        "observed_response": r.get("observed_response"),
        "blocked_by_safety": r.get("blocked_by_safety"),
        "evidence_file": r.get("evidence_file"),
    } for r in all_can])

    write_sheet(wb, "表5_模型对比", [
        "目标", "artifact_run_id", "variant_id", "variant_label", "elapsed_seconds", "executed_poc_count", "finding_count", "llm_call_count", "total_tokens", "agent_execution_coverage_vs_global", "reflection_reentry_count", "success",
    ], all_model)

    try:
        from compare_global_vs_agent import TABLE8_FOOTNOTE, TABLE8_PRIMARY_METRIC_NOTE
    except Exception:
        TABLE8_FOOTNOTE = "主指标 gt_recall；Agent 为定向验证。"
        TABLE8_PRIMARY_METRIC_NOTE = TABLE8_FOOTNOTE

    write_sheet(wb, "表8_指标说明", ["项", "说明"], [
        {"项": "主指标", "说明": "gt_recall（命中数/gt_positive_count）；论文主列用 paper_primary_recall"},
        {"项": "Agent 定位", "说明": "定向验证：仅对侦察/Global 已检出攻击面执行少量 PoC，不与 Global 全量 recon 比召回"},
        {"项": "效率", "说明": "poc_reduction_percent = 较 Global 少跑的 PoC 比例"},
        {"项": "参考列", "说明": "agent_finding_recall_vs_global 以 Global 检出为分母，不代表真值，勿与 gt_recall 混用"},
        {"项": "脚注", "说明": TABLE8_FOOTNOTE},
        {"项": "解读", "说明": TABLE8_PRIMARY_METRIC_NOTE},
    ])

    write_sheet(wb, "表8_Global与Agent对比", [
        "target_id", "artifact_run_id", "variant_id",
        "gt_positive_count", "gt_hit_count", "gt_recall", "paper_primary_recall", "agent_efficiency_score",
        "global_gt_recall", "global_vulnerable_count", "agent_finding_count", "agent_findings_per_executed_poc",
        "agent_executed_poc_count", "poc_reduction_percent", "agent_elapsed_seconds", "global_elapsed_seconds",
        "time_ratio_global_over_agent", "agent_total_tokens",
        "finding_overlap_with_global", "agent_finding_recall_vs_global", "agent_finding_precision_vs_global",
        "same_conclusion_note",
    ], all_compare)

    write_sheet(wb, "安全拦截明细", ["目标", "PoC编号", "执行状态", "是否拦截", "证据文件"], all_blocked)

    write_sheet(wb, "典型案例", ["case_id", "case_name", "target_id", "attack_surface", "expected_evidence"], typical_cases)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(json.dumps({"output": str(args.output), "targets": discover_targets(args.evidence_root)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
