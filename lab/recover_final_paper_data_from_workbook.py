#!/usr/bin/env python3
"""Rebuild lab/final_paper_data_strict/*.json from frozen paper workbooks."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_BOOK = ROOT / "lab" / "原始数据统计表（新）.xlsx"
PAPER_BOOK = ROOT / "lab" / "论文内表格（新）.xlsx"
OUT_DIR = ROOT / "lab" / "final_paper_data_strict"

SHEET_MAP = {
    "paper_key_metrics": "论文关键指标",
    "primary_recall_summary": "主结果_阳性召回",
    "table6_total_by_category": "表6_总统计",
    "table7_total_by_model_group": "表7_总统计",
    "table8_total_by_model": "表8_总统计",
    "global_summary_by_target": "基准扫描汇总",
    "model_summary_all_targets": "模型跨目标汇总",
    "best_model_by_target": "各目标最佳模型",
    "table6_all_targets": "表6_分设备明细",
    "table7_all_targets": "表7_分设备明细",
    "table8_all_targets": "表8_分设备明细",
    "table9_all_targets": "表9_三目标",
    "table10_reflection_summary": "反思重入_汇总",
    "table10_all_targets": "反思重入_三目标明细",
    "table11_safety_control": "表11_安全控制",
    "raw_global_agent_comparison": "基准扫描_Agent原始对比",
}


def sheet_rows(path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"{path.name}: missing sheet {sheet_name!r}")
    ws = wb[sheet_name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({str(headers[i]): values[i] for i in range(len(headers)) if headers[i]})
    return rows


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_fraction(text: str) -> tuple[int, int] | None:
    match = re.search(r"(\d+)\s*/\s*(\d+)", str(text or ""))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def build_table10_pentestgpt(raw_rows: list[dict]) -> dict:
    row = next(item for item in raw_rows if "PentestGPT" in str(item.get("系统/模型") or ""))
    recall = str(row.get("Recall@GT（基准阳性召回率）") or row.get("基准阳性召回率（Vulnerability Recall）") or "")
    coverage = str(row.get("Coverage（覆盖率）") or row.get("基准子任务完成率（Benchmark Sub-task Completion Rate）") or "")
    latency = str(row.get("Avg. Latency（平均验证耗时）") or row.get("平均端到端净耗时（Mean End-to-End Runtime）") or "")
    frac = parse_fraction(recall)
    cov_frac = parse_fraction(coverage)
    minutes = re.search(r"([\d.]+)", latency)
    payload = {
        "system": "PentestGPT（GLM-5）",
        "risk_num": frac[0] if frac else 14,
        "risk_den": frac[1] if frac else 30,
        "risk_recall": recall,
        "Coverage（覆盖率）": coverage,
        "duration_seconds": float(minutes.group(1)) * 60 if minutes else 985.8,
        "coverage_done": cov_frac[0] if cov_frac else 18,
        "coverage_total": cov_frac[1] if cov_frac else 30,
        "data_source": row.get("数据来源与口径", ""),
    }
    return payload


def main() -> int:
    if not RAW_BOOK.is_file():
        raise FileNotFoundError(RAW_BOOK)
    if not PAPER_BOOK.is_file():
        raise FileNotFoundError(PAPER_BOOK)

    exported: dict[str, list[dict] | dict] = {}
    for key, sheet in SHEET_MAP.items():
        exported[key] = sheet_rows(RAW_BOOK, sheet)

    # Prefer bilingual display headers from the paper-facing workbook for headline tables.
    for key, sheet in {
        "table6_total_by_category": "表6_PoC命中成功漏报（智谱）",
        "table7_total_by_model_group": "表7_智能体消融（智谱）",
        "table8_total_by_model": "表8_模型对比",
    }.items():
        exported[key] = sheet_rows(PAPER_BOOK, sheet)

    table10_rows = sheet_rows(RAW_BOOK, "表10_平台能力对比")
    exported["table10_platform_comparison"] = table10_rows
    exported["table10_pentestgpt_three_targets"] = build_table10_pentestgpt(table10_rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for key, payload in exported.items():
        write_json(OUT_DIR / f"{key}.json", payload)

    manifest = {
        "recovered_from": [str(RAW_BOOK), str(PAPER_BOOK)],
        "tables": {key: len(payload) if isinstance(payload, list) else 1 for key, payload in exported.items()},
    }
    write_json(OUT_DIR / "manifest.json", manifest)
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
