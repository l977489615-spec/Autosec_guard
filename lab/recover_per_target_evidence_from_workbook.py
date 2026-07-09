#!/usr/bin/env python3
"""Rebuild per-target evidence skeletons from frozen paper workbooks.

Recovers what can be inferred without the original agent report JSON files:
- lab/evidence/<TARGET>/selected_agent_runs.json
- lab/evidence/<TARGET>/scan_results.json (baseline positives only)
- lab/ground_truth/<TARGET>.json (IVI-01 / REAL-CAR)
- lab/final_paper_data_strict/raw_scan_rows.json
- lab/evidence/MISSING_AGENT_REPORTS.json (index of files still absent on disk)
"""
from __future__ import annotations

import argparse
import ast
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW_BOOK = ROOT / "lab" / "原始数据统计表（新）.xlsx"
EVIDENCE_ROOT = ROOT / "lab" / "evidence"
GT_DIR = ROOT / "lab" / "ground_truth"
OUT_STRICT = ROOT / "lab" / "final_paper_data_strict"
TARGETS = ("MOCK-LOCAL", "IVI-01", "REAL-CAR")


def read_json(path: Path, default=None):
    if not path.is_file():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_poc_list(value) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item) for item in value]
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = ast.literal_eval(text)
    except Exception:
        return [text]
    if isinstance(parsed, list):
        return [str(item) for item in parsed]
    return [str(parsed)]


def load_comparison_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["基准扫描_Agent原始对比"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({str(headers[i]): values[i] for i in range(len(headers)) if headers[i]})
    return rows


def baseline_union(rows: list[dict], target_id: str) -> list[str]:
    union: set[str] = set()
    for row in rows:
        if str(row.get("target_id") or "") != target_id:
            continue
        union.update(parse_poc_list(row.get("baseline_overlap_pocs")))
    return sorted(union)


def primary_row(rows: list[dict], target_id: str, variant_id: str = "ZHIPU") -> dict:
    for row in rows:
        if str(row.get("target_id") or "") == target_id and str(row.get("variant_id") or "") == variant_id:
            return row
    raise KeyError(f"missing comparison row for {target_id}/{variant_id}")


def infer_ground_truth_positives(row: dict) -> list[str]:
    hits = parse_poc_list(row.get("finding_overlap_pocs"))
    baseline = parse_poc_list(row.get("baseline_overlap_pocs"))
    gt_count = int(row.get("gt_positive_count") or len(hits) or 0)
    positives = list(hits)
    if len(positives) >= gt_count:
        return sorted(positives[:gt_count])
    for poc in baseline:
        if poc not in positives:
            positives.append(poc)
        if len(positives) >= gt_count:
            break
    return sorted(positives[:gt_count])


def scan_row(target_id: str, poc_file: str) -> dict:
    category = poc_file.split("/", 1)[0] if "/" in poc_file else "network"
    return {
        "target_id": target_id,
        "poc_file": poc_file,
        "category": category,
        "status": "completed",
        "elapsed_seconds": 0.0,
        "vulnerable": True,
        "evidence": "reconstructed from frozen workbook baseline_overlap_pocs",
        "reconstructed": True,
    }


def build_selected_manifest(target_id: str, rows: list[dict]) -> dict:
    agent_files: list[dict] = []
    missing: list[dict] = []
    for row in rows:
        if str(row.get("target_id") or "") != target_id:
            continue
        report_file = str(row.get("report_file") or "").strip()
        if not report_file:
            continue
        rel = report_file.split(f"/{target_id}/", 1)[-1]
        if not rel.startswith("agent_runs/"):
            rel = f"agent_runs/{Path(report_file).name}"
        abs_path = ROOT / report_file
        entry = {
            "variant_id": str(row.get("variant_id") or ""),
            "path": rel,
            "source_report_file": report_file,
            "exists_on_disk": abs_path.is_file(),
        }
        agent_files.append(entry)
        if not abs_path.is_file():
            missing.append(entry)
    return {
        "target_id": target_id,
        "schema_version": 1,
        "recovered_from": str(RAW_BOOK),
        "note": "路径来自冻结工作簿；原始 Agent 报告 JSON 需从备份或重跑实验恢复。",
        "agent_run_files": [item["path"] for item in agent_files],
        "variants": agent_files,
        "missing_reports": missing,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Recover per-target evidence skeletons from workbook.")
    parser.add_argument("--workbook", type=Path, default=RAW_BOOK)
    parser.add_argument("--evidence-root", type=Path, default=EVIDENCE_ROOT)
    parser.add_argument("--gt-dir", type=Path, default=GT_DIR)
    parser.add_argument("--strict-dir", type=Path, default=OUT_STRICT)
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise FileNotFoundError(args.workbook)

    rows = load_comparison_rows(args.workbook)
    all_missing: list[dict] = []
    raw_scan_rows: list[dict] = []

    for target_id in TARGETS:
        target_rows = [row for row in rows if str(row.get("target_id") or "") == target_id]
        manifest = build_selected_manifest(target_id, target_rows)
        target_dir = args.evidence_root / target_id
        write_json(target_dir / "selected_agent_runs.json", manifest)
        all_missing.extend(manifest["missing_reports"])

        positives = baseline_union(rows, target_id)
        scan_rows = [scan_row(target_id, poc_file) for poc_file in positives]
        write_json(target_dir / "scan_results.json", scan_rows)
        raw_scan_rows.extend(scan_rows)

        if target_id == "MOCK-LOCAL" and (args.gt_dir / "MOCK-LOCAL.json").is_file():
            continue

        zhipu = primary_row(rows, target_id)
        gt_payload = {
            "target_id": target_id,
            "description": f"{target_id} ground truth（由冻结工作簿重建）",
            "positive_pocs": infer_ground_truth_positives(zhipu),
            "negative_pocs": [],
            "notes": "由 基准扫描_Agent原始对比 中 ZHIPU 行的 finding_overlap + gt_positive_count 推断；REAL-CAR 末 4 项为 baseline 补全，需人工复核。",
            "scan_confirmed_count": len(parse_poc_list(zhipu.get("baseline_overlap_pocs"))),
            "reconstructed": True,
        }
        write_json(args.gt_dir / f"{target_id}.json", gt_payload)

    write_json(args.strict_dir / "raw_scan_rows.json", raw_scan_rows)
    write_json(
        args.evidence_root / "MISSING_AGENT_REPORTS.json",
        {
            "recovered_from": str(args.workbook),
            "missing_count": len(all_missing),
            "reports": all_missing,
            "recovery_options": [
                "Time Machine / 外置备份盘搜索 AGENT-ZHIPU_20260604_034605.json",
                "按 selected_agent_runs.json 中的路径重跑四模型 Agent 实验",
            ],
        },
    )

    summary = {
        "targets": list(TARGETS),
        "scan_rows": len(raw_scan_rows),
        "missing_agent_reports": len(all_missing),
        "ground_truth_written": [f"{target}.json" for target in TARGETS if target != "MOCK-LOCAL" or not (args.gt_dir / "MOCK-LOCAL.json").exists()],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
