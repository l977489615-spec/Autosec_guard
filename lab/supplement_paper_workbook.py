#!/usr/bin/env python3
"""Supplement Excel workbooks with paper-aligned tables — do not modify thesis docx."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from metric_definitions import AVG_LATENCY, COVERAGE, HITS_AT_GT_FRACTION, MISS_RATE, RECALL_AT_GT
from paper_table_rows import ablation_latency_minutes, build_paper_table_rows, load_json
from sync_paper_table_values import (
    DEFAULT_PAPER_BOOK,
    SUMMARY_BOOK,
    TABLE7_SHEET,
    restore_summary_tables_from_paper_book,
    resolve_sheet_name,
    update_paper_workbook,
)

PAPER_SHEETS = {
    "论文表2_PoC验证闭环": ("poc", [22, 14, 42]),
    "论文表3_智能体消融": ("ablation", [18, 16, 14, 12, 16]),
    "论文表4_规划策略": ("strategy", [18, 34, 14, 14, 12]),
    "论文表5_基线对比": ("baseline", [20, 16, 14, 12, 16]),
    "论文表6_模型对比": ("model", [22, 16, 14, 12, 18, 14]),
}

HEADER_FILL = PatternFill("solid", fgColor="CFE8F6")
HEADER_FONT = Font(name="宋体", bold=True, color="000000")
THIN = Side(style="thin", color="808080")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_table(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = GRID
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = ws.dimensions


def replace_or_create_sheet(wb, name: str, rows: list[list], widths: list[int], *, index: int | None = None) -> None:
    if name in wb.sheetnames:
        old = wb[name]
        idx = wb.sheetnames.index(name)
        wb.remove(old)
    else:
        idx = index if index is not None else len(wb.sheetnames)
    ws = wb.create_sheet(name, idx)
    for row in rows:
        ws.append(row)
    style_table(ws, widths)


def patch_table7_latency(ws) -> int:
    headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    latency_col = headers.get(AVG_LATENCY)
    group_col = headers.get("组别")
    if not latency_col or not group_col:
        return 0
    zhipu_latency = next(
        row["Avg. Latency（平均验证耗时）"]
        for row in load_json("table8_total_by_model.json")
        if row.get("variant_id") == "ZHIPU"
    )
    updates = {
        "A": ablation_latency_minutes("A"),
        "B": "-",
        "C": "-",
        "D": zhipu_latency,
    }
    count = 0
    for row_idx in range(2, ws.max_row + 1):
        group = str(ws.cell(row_idx, group_col).value or "").strip()
        if group not in updates:
            continue
        value = updates[group] or "-"
        if ws.cell(row_idx, latency_col).value != value:
            ws.cell(row_idx, latency_col, value)
            count += 1
    return count


def patch_table10_miss_rate(ws) -> int:
    headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    if MISS_RATE in headers:
        return 0
    coverage_col = headers.get(COVERAGE)
    if not coverage_col:
        return 0
    insert_at = coverage_col + 1
    ws.insert_cols(insert_at, 1)
    ws.cell(1, insert_at, MISS_RATE)
    label_col = headers.get("系统/模型") or headers.get("方案") or 1
    pgpt = load_json("table10_pentestgpt_three_targets.json")
    pgpt_miss = f"{(int(pgpt['risk_den']) - int(pgpt['risk_num'])) / int(pgpt['risk_den']) * 100:.1f}%（{int(pgpt['risk_den']) - int(pgpt['risk_num'])}/{pgpt['risk_den']}）"
    platform_miss = {
        "PentestGPT（GLM-5）": pgpt_miss,
    }
    zhipu_miss = next(
        row["Miss Rate（漏报率）"]
        for row in load_json("table8_total_by_model.json")
        if row.get("variant_id") == "ZHIPU"
    )
    for row_idx in range(2, ws.max_row + 1):
        label = str(ws.cell(row_idx, label_col).value or "")
        if "PentestGPT" in label:
            ws.cell(row_idx, insert_at, platform_miss)
        elif "智谱" in label or "GLM-5" in label:
            ws.cell(row_idx, insert_at, zhipu_miss)
        elif "GPT" in label and "Pentest" not in label:
            row = next(r for r in load_json("table8_total_by_model.json") if r.get("variant_id") == "GPT")
            ws.cell(row_idx, insert_at, row["Miss Rate（漏报率）"])
        else:
            ws.cell(row_idx, insert_at, "-")
    return ws.max_row - 1


def patch_reflection_sheet(ws) -> None:
    t7_total = next(row for row in load_json("table7_total_by_model_group.json") if row.get("组别") == "D")
    rows = [
        ["指标", "数值", "数据来源"],
        ["首次执行失败步骤数", 8, "三目标Agent报告聚合"],
        ["触发反思次数", 14, "三目标Agent报告聚合"],
        ["定向重跑次数", 14, "三目标Agent报告聚合"],
        ["补充侦察/补证问题数", 20, "三目标Agent报告聚合"],
        ["补证成功数量", t7_total.get("Agent命中阳性数", ""), "比较结果聚合"],
        ["最终漏洞检出率", t7_total.get(RECALL_AT_GT, ""), t7_total.get(HITS_AT_GT_FRACTION, "")],
    ]
    while ws.max_row > 0:
        ws.delete_rows(1)
    for row in rows:
        ws.append(row)
    style_table(ws, [24, 15, 30])


def supplement_workbook(path: Path) -> dict:
    rows_by_key = build_paper_table_rows()
    wb = load_workbook(path)

    for sheet_name, (key, widths) in PAPER_SHEETS.items():
        replace_or_create_sheet(wb, sheet_name, rows_by_key[key], widths, index=2)

    table7_sheet = resolve_sheet_name(wb, TABLE7_SHEET)
    table7_patched = 0
    if table7_sheet:
        table7_patched = patch_table7_latency(wb[table7_sheet])

    table10_patched = 0
    if "表10_平台能力对比" in wb.sheetnames:
        table10_patched = patch_table10_miss_rate(wb["表10_平台能力对比"])

    reflection_sheet = None
    for candidate in ("表10_反思重入", "反思重入实验"):
        if candidate in wb.sheetnames:
            reflection_sheet = candidate
            break
    if reflection_sheet:
        patch_reflection_sheet(wb[reflection_sheet])

    wb.save(path)
    return {
        "path": str(path),
        "paper_sheets": list(PAPER_SHEETS),
        "table7_latency_rows": table7_patched,
        "table10_miss_rate_rows": table10_patched,
        "reflection_updated": reflection_sheet is not None,
    }


def main() -> None:
    targets = [DEFAULT_PAPER_BOOK, SUMMARY_BOOK]
    results = []
    for book in targets:
        if not book.is_file():
            continue
        if book == DEFAULT_PAPER_BOOK:
            update_paper_workbook(book)
        result = supplement_workbook(book)
        results.append(result)
    if DEFAULT_PAPER_BOOK.is_file():
        restore_summary_tables_from_paper_book(DEFAULT_PAPER_BOOK)
        if SUMMARY_BOOK.is_file():
            supplement_workbook(SUMMARY_BOOK)
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
