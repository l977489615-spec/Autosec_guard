#!/usr/bin/env python3
"""Update table 6/7/8 data cells in-place, preserving existing workbook layout and styles."""

from __future__ import annotations

import json
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from l3_evidence_rates import (
    PAPER_EVIDENCE_HEADER,
    enrich_table6_rows,
    enrich_table7_rows,
    enrich_table8_rows,
    table10_evidence_rates,
)
from metric_definitions import (
    COVERAGE,
    COVERED_TASKS_FRACTION,
    EVIDENCE_COMPLETENESS,
    GT_EXPOSURE_FRACTION,
    GT_EXPOSURE_RATE,
    HITS_AT_GT_FRACTION,
    LEGACY_HEADER_ALIASES,
    MEAN_E2E_RUNTIME,
    MISSED_GT_FRACTION,
    MISS_RATE,
    RECALL_AT_GT,
)
from paper_metric_names import JSON_LATENCY_KEY, JSON_MISS_KEY, JSON_RECALL_KEY, JSON_SUBTASK_KEY


ROOT = Path(__file__).resolve().parents[1]
STRICT_DATA = ROOT / "lab" / "final_paper_data_strict"
DEFAULT_PAPER_BOOK = Path("/Users/queen/Desktop/ICV_POC_research/论文/论文内表格（新）.xlsx")
SUMMARY_BOOK = ROOT / "lab" / "论文实验数据汇总.xlsx"

TABLE6_SHEET = "表6_PoC命中成功漏报"
TABLE7_SHEET = "表7_智能体消融"
TABLE8_SHEET = "表8_模型对比"
TABLE10_SHEET = "表10_平台能力对比"


def resolve_sheet_name(wb, canonical: str) -> str | None:
    if canonical in wb.sheetnames:
        return canonical
    for name in wb.sheetnames:
        if name.startswith(canonical):
            return name
    return None


def workbook_has_strict_tables(path: Path) -> bool:
    if not path.is_file():
        return False
    wb = load_workbook(path, read_only=True)
    return bool(resolve_sheet_name(wb, TABLE7_SHEET) and resolve_sheet_name(wb, TABLE8_SHEET))

REMOVED_HEADERS = {
    "宏平均成功率（Macro Success Rate）",
    "Macro Success Rate（宏平均成功率）",
    "Evidence Archive Rate（证据归档率）",
    "有效证据数/已执行数",
    "Archived/Executed（归档/已执行数）",
    "Audited/Executed（可审计/已执行数）",
    "有效证据数",
    "任务完成率",
    "Task Completion Rate（任务完成率）",
    "完成任务数/任务总数",
    "Completed/|T|（完成/任务总数）",
    "执行完成率",
    "Execution Completion Rate（执行完成率）",
    "PoC选择召回率",
    "PoC Selection Recall（PoC选择召回率）",
    "已执行PoC数",
    "平均任务耗时",
    "Per-target Coverage（分目标覆盖率）",
    "分设备覆盖项数/分设备任务总数",
    "分设备覆盖率",
    "分设备覆盖项数",
    "分设备任务总数",
}

SHEET_HEADER_RENAMES = {
    **LEGACY_HEADER_ALIASES,
    "Global阳性PoC数": "基准阳性PoC数",
    "命中率": GT_EXPOSURE_RATE,
    "执行 PoC 总数": "已执行数量",
    "发现数合计": "Agent命中阳性数",
}

TABLE6_FIELDS = {
    "类别": "类别",
    "PoC 数量": "PoC 数量",
    "已执行数量": "已执行数量",
    "基准阳性PoC数": "基准阳性PoC数",
    "Agent命中阳性数": "Agent命中阳性数",
    "漏报数": "漏报数",
    RECALL_AT_GT: JSON_RECALL_KEY,
    HITS_AT_GT_FRACTION: HITS_AT_GT_FRACTION,
    COVERAGE: JSON_SUBTASK_KEY,
    COVERED_TASKS_FRACTION: COVERED_TASKS_FRACTION,
    MISS_RATE: "漏报率",
    MISSED_GT_FRACTION: "漏报数/基准阳性PoC数",
    # legacy JSON keys still accepted via field values below
    GT_EXPOSURE_RATE: GT_EXPOSURE_RATE,
    GT_EXPOSURE_FRACTION: GT_EXPOSURE_FRACTION,
}

TABLE7_FIELDS = {
    "组别": "组别",
    "系统配置": "系统配置",
    HITS_AT_GT_FRACTION: HITS_AT_GT_FRACTION,
    RECALL_AT_GT: JSON_RECALL_KEY,
    COVERAGE: JSON_SUBTASK_KEY,
    COVERED_TASKS_FRACTION: COVERED_TASKS_FRACTION,
    MISSED_GT_FRACTION: MISSED_GT_FRACTION,
    MISS_RATE: JSON_MISS_KEY,
    "平均人工干预次数": "平均人工干预次数",
    MEAN_E2E_RUNTIME: JSON_LATENCY_KEY,
    PAPER_EVIDENCE_HEADER: PAPER_EVIDENCE_HEADER,
    "data_source": "data_source",
}

TABLE8_FIELDS = {
    "模型": "模型",
    "类型": "类型",
    RECALL_AT_GT: JSON_RECALL_KEY,
    COVERAGE: JSON_SUBTASK_KEY,
    COVERED_TASKS_FRACTION: COVERED_TASKS_FRACTION,
    MISSED_GT_FRACTION: MISSED_GT_FRACTION,
    MISS_RATE: JSON_MISS_KEY,
    MEAN_E2E_RUNTIME: JSON_LATENCY_KEY,
    "高风险误触发次数": "高风险误触发次数",
    "目标数量": "目标数量",
    HITS_AT_GT_FRACTION: HITS_AT_GT_FRACTION,
    "已执行数量": "已执行数量",
    "Agent命中阳性数": "Agent命中阳性数",
    "Total Tokens 合计": "Total Tokens 合计",
    "平均每目标 Tokens": "平均每目标 Tokens",
    "总耗时": "总耗时",
    "variant_id": "variant_id",
    PAPER_EVIDENCE_HEADER: PAPER_EVIDENCE_HEADER,
    "data_source": "data_source",
}

TABLE10_FIELDS = {
    "系统/模型": "系统/模型",
    PAPER_EVIDENCE_HEADER: PAPER_EVIDENCE_HEADER,
}


def strict_rows(name: str) -> list[dict]:
    payload = json.loads((STRICT_DATA / f"{name}.json").read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"严格汇总数据格式错误: {name}")
    return payload


def header_columns(ws: Worksheet) -> dict[str, int]:
    columns: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header is not None and str(header).strip():
            columns[str(header).strip()] = col
    return columns


def rename_headers(ws: Worksheet, rename_map: dict[str, str]) -> None:
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header is None:
            continue
        text = str(header).strip()
        if text in rename_map:
            target = rename_map[text]
            if target != "__REMOVE__":
                ws.cell(1, col, target)


def remove_columns(ws: Worksheet, headers: set[str]) -> None:
    columns = header_columns(ws)
    for header in sorted(headers, key=lambda item: columns.get(item, 0), reverse=True):
        col = columns.get(header)
        if col:
            ws.delete_cols(col)
            columns = header_columns(ws)


def remove_duplicate_columns(ws: Worksheet, header: str, *, keep: int = 1) -> None:
    matches = [
        col
        for col in range(1, ws.max_column + 1)
        if str(ws.cell(1, col).value or "").strip() == header
    ]
    for col in reversed(matches[keep:]):
        ws.delete_cols(col)


def update_rows_by_key(
    ws: Worksheet,
    rows: list[dict],
    *,
    key_header: str,
    field_map: dict[str, str],
    json_key: str,
) -> int:
    columns = header_columns(ws)
    key_col = columns.get(key_header)
    if not key_col:
        raise ValueError(f"工作表 {ws.title} 缺少关键列: {key_header}")

    updated = 0
    for item in rows:
        key_value = item.get(json_key)
        if key_value is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row_idx, key_col).value != key_value:
                continue
            for excel_header, json_field in field_map.items():
                col = columns.get(excel_header)
                if col is None:
                    continue
                if json_field in item:
                    value = item[json_field]
                    if value in ("", None) and excel_header == MEAN_E2E_RUNTIME:
                        value = "-"
                    ws.cell(row_idx, col, value)
            updated += 1
            break
    return updated


def copy_cell_style(source_cell, target_cell) -> None:
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def copy_worksheet(source_ws: Worksheet, target_wb, name: str, index: int) -> Worksheet:
    if name in target_wb.sheetnames:
        target_wb.remove(target_wb[name])
    target_ws = target_wb.create_sheet(name, index)
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(cell.row, cell.column, cell.value)
            copy_cell_style(cell, target_cell)
    for col, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col].width = dim.width
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    return target_ws


def ensure_column_after(ws: Worksheet, after_header: str, new_header: str) -> int | None:
    columns = header_columns(ws)
    if new_header in columns:
        return columns[new_header]
    after_col = columns.get(after_header)
    if not after_col:
        return None
    insert_at = after_col + 1
    ws.insert_cols(insert_at, 1)
    anchor_header_cell = ws.cell(1, after_col)
    new_header_cell = ws.cell(1, insert_at, new_header)
    copy_cell_style(anchor_header_cell, new_header_cell)
    for row_idx in range(2, ws.max_row + 1):
        copy_cell_style(ws.cell(row_idx, after_col), ws.cell(row_idx, insert_at))
    return insert_at


def update_paper_workbook(path: Path) -> dict[str, int]:
    wb = load_workbook(path)
    table7_sheet = resolve_sheet_name(wb, TABLE7_SHEET)
    table8_sheet = resolve_sheet_name(wb, TABLE8_SHEET)
    if not table7_sheet or not table8_sheet:
        raise FileNotFoundError(f"{path} 缺少 {TABLE7_SHEET} 或 {TABLE8_SHEET}")

    table6_count = 0
    table6_sheet = resolve_sheet_name(wb, TABLE6_SHEET)
    if table6_sheet:
        ws6 = wb[table6_sheet]
        remove_columns(ws6, REMOVED_HEADERS)
        rename_headers(ws6, SHEET_HEADER_RENAMES)
        ensure_column_after(ws6, RECALL_AT_GT, HITS_AT_GT_FRACTION)
        ensure_column_after(ws6, HITS_AT_GT_FRACTION, COVERAGE)
        ensure_column_after(ws6, COVERAGE, COVERED_TASKS_FRACTION)
        ensure_column_after(ws6, MISS_RATE, MISSED_GT_FRACTION)
        ensure_column_after(ws6, GT_EXPOSURE_RATE, GT_EXPOSURE_FRACTION)
        remove_columns(ws6, {PAPER_EVIDENCE_HEADER})
        table6_count = update_rows_by_key(
            ws6,
            enrich_table6_rows(strict_rows("table6_total_by_category")),
            key_header="类别",
            field_map=TABLE6_FIELDS,
            json_key="类别",
        )

    ws7 = wb[table7_sheet]
    ws8 = wb[table8_sheet]
    remove_columns(ws7, REMOVED_HEADERS)
    remove_columns(ws8, REMOVED_HEADERS)
    rename_headers(ws7, SHEET_HEADER_RENAMES)
    rename_headers(ws8, SHEET_HEADER_RENAMES)
    remove_columns(ws7, {"宏平均成功率（Macro Success Rate）", "Macro Success Rate（宏平均成功率）"})
    remove_columns(ws8, {"宏平均成功率（Macro Success Rate）", "Macro Success Rate（宏平均成功率）"})
    ensure_column_after(ws7, RECALL_AT_GT, COVERAGE)
    ensure_column_after(ws7, COVERAGE, COVERED_TASKS_FRACTION)
    remove_duplicate_columns(ws7, COVERAGE)
    ensure_column_after(ws7, MISS_RATE, MISSED_GT_FRACTION)
    ensure_column_after(ws7, COVERAGE, PAPER_EVIDENCE_HEADER)
    ensure_column_after(ws8, RECALL_AT_GT, COVERAGE)
    ensure_column_after(ws8, COVERAGE, COVERED_TASKS_FRACTION)
    remove_duplicate_columns(ws8, COVERAGE)
    ensure_column_after(ws8, MISS_RATE, MISSED_GT_FRACTION)
    ensure_column_after(ws8, COVERAGE, PAPER_EVIDENCE_HEADER)

    table7_count = update_rows_by_key(
        ws7,
        enrich_table7_rows(strict_rows("table7_total_by_model_group")),
        key_header="组别",
        field_map=TABLE7_FIELDS,
        json_key="组别",
    )
    table8_count = update_rows_by_key(
        ws8,
        enrich_table8_rows(strict_rows("table8_total_by_model")),
        key_header="模型",
        field_map=TABLE8_FIELDS,
        json_key="模型",
    )

    table10_count = 0
    table10_sheet = resolve_sheet_name(wb, TABLE10_SHEET)
    if table10_sheet:
        ws10 = wb[table10_sheet]
        rename_headers(ws10, SHEET_HEADER_RENAMES)
        remove_columns(ws10, REMOVED_HEADERS)
        remove_columns(ws10, {"宏平均成功率（Macro Success Rate）", "Macro Success Rate（宏平均成功率）"})
        ensure_column_after(ws10, COVERAGE, PAPER_EVIDENCE_HEADER)
        remove_duplicate_columns(ws10, PAPER_EVIDENCE_HEADER)
        table10_rows = [
            {"系统/模型": label, PAPER_EVIDENCE_HEADER: value}
            for label, value in table10_evidence_rates().items()
        ]
        table10_count = update_rows_by_key(
            ws10,
            table10_rows,
            key_header="系统/模型",
            field_map=TABLE10_FIELDS,
            json_key="系统/模型",
        )

    wb.save(path)
    return {
        "table6_rows": table6_count,
        "table7_rows": table7_count,
        "table8_rows": table8_count,
        "table10_rows": table10_count,
        "path": str(path),
    }


def restore_summary_tables_from_paper_book(paper_book: Path = DEFAULT_PAPER_BOOK) -> None:
    if not paper_book.is_file():
        return
    try:
        paper_wb = load_workbook(paper_book)
        summary_wb = load_workbook(SUMMARY_BOOK)
        for canonical in (TABLE6_SHEET, TABLE7_SHEET, TABLE8_SHEET, TABLE10_SHEET):
            paper_sheet = resolve_sheet_name(paper_wb, canonical)
            if not paper_sheet:
                continue
            index = (
                summary_wb.sheetnames.index(canonical)
                if canonical in summary_wb.sheetnames
                else len(summary_wb.sheetnames)
            )
            copy_worksheet(paper_wb[paper_sheet], summary_wb, canonical, index)
        summary_wb.save(SUMMARY_BOOK)
    except Exception:
        shutil.copy2(paper_book, SUMMARY_BOOK)


def main() -> None:
    paper_book = DEFAULT_PAPER_BOOK
    if not paper_book.is_file():
        raise SystemExit(f"未找到论文表格: {paper_book}")
    result = update_paper_workbook(paper_book)
    restore_summary_tables_from_paper_book(paper_book)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
