#!/usr/bin/env python3
"""Update the paper workbooks from frozen platform data and PentestGPT results."""

from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from metric_definitions import (
    AVG_LATENCY,
    COVERAGE,
    HITS_AT_GT_FRACTION,
    MACRO_SUCCESS_RATE,
    RECALL_AT_GT,
    rate_display,
)
from sync_paper_table_values import (
    TABLE7_SHEET,
    TABLE8_SHEET,
    restore_summary_tables_from_paper_book,
    update_paper_workbook,
    workbook_has_strict_tables,
)
import sys

ROOT = Path(__file__).resolve().parents[1]
LAB = ROOT / "lab"
PGPT_DIR = LAB / "pentestgpt_id4"
if str(PGPT_DIR) not in sys.path:
    sys.path.insert(0, str(PGPT_DIR))
from pentestgpt_platform_scoring import (  # noqa: E402
    available_targets as pgpt_available_targets,
    score_pentestgpt_report_set,
)
from table10_platform_scoring import (  # noqa: E402
    TABLE10_TARGETS,
    score_platform_report_set,
    write_manifest_cache,
)

TABLE10_TARGETS_NOTE = "MOCK-LOCAL + IVI-01（192.168.31.158）+ REAL-CAR"
TOTAL_BOOK = (
    LAB / "论文实验数据_总.xlsx"
    if (LAB / "论文实验数据_总.xlsx").is_file()
    else LAB / "实验数据_总.xlsx"
)
SUMMARY_BOOK = LAB / "论文实验数据汇总.xlsx"
PAPER_BOOKS = [
    Path("/Users/queen/Desktop/ICV_POC_research/论文/原始数据统计表（新）.xlsx"),
    Path("/Users/queen/Desktop/ICV_POC_research/论文/论文内表格（新）.xlsx"),
]
TABLE10_PLATFORM_SHEET = "表10_平台能力对比"
TABLE10_PLATFORM_WIDTHS = [22, 18, 14, 16, 14, 14, 16, 12, 34]
PGPT_RESULTS = LAB / "pentestgpt_id4" / "results" / "pentestgpt_per_run.json"
STRICT_DATA = LAB / "final_paper_data_strict"

VARIANT_LABELS = {
    "GPT": "OpenAI GPT-5.4-mini",
    "ZHIPU": "智谱 GLM-5",
    "DEEPSEEK": "DeepSeek v4 pro",
    "QWEN-MAX": "千问 qwen-max（质量）",
}

HEADER_FILL = PatternFill("solid", fgColor="CFE8F6")
HEADER_FONT = Font(name="宋体", bold=True, color="000000")
BODY_FONT = Font(name="宋体", size=10)
THIN = Side(style="thin", color="808080")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def strict_rows(name: str) -> list[dict]:
    path = STRICT_DATA / f"{name}.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"严格汇总数据格式错误: {path}")
    return payload


def style_table(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = GRID
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = ws.dimensions


def replace_sheet(wb, name: str, rows: list[list], widths: list[int], index=None):
    if name in wb.sheetnames:
        old = wb[name]
        index = wb.sheetnames.index(name) if index is None else index
        wb.remove(old)
    ws = wb.create_sheet(name, index if index is not None else len(wb.sheetnames))
    for row in rows:
        ws.append(row)
    style_table(ws, widths)
    return ws


def rename_if_present(wb, old: str, new: str) -> None:
    if old in wb.sheetnames and new not in wb.sheetnames:
        wb[old].title = new


def update_pass_columns(wb) -> None:
    for sheet_name in ("表8_总统计", "表8_分设备明细", "表8_模型对比"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}
        pass_col = headers.get("Pass@5")
        success_col = headers.get("整体任务成功率")
        if not pass_col or not success_col:
            continue
        ws.cell(1, pass_col, RECALL_AT_GT)
        for row in range(2, ws.max_row + 1):
            ws.cell(row, pass_col, ws.cell(row, success_col).value)


def update_library_coverage(ws) -> None:
    counts = {
        "侦察类": 8,
        "网络服务类": 15,
        "应用安全类": 46,
        "系统配置类": 12,
        "第三方组件类": 5,
        "车内协议类": 13,
        "无线与外设类": 37,
        "高级攻击类": 10,
    }
    headers = {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}
    for row in range(2, ws.max_row + 1):
        category = str(ws.cell(row, headers["攻击面类别"]).value or "")
        if category in counts:
            ws.cell(row, headers["PoC 数量"], counts[category])
            if "data_source" in headers:
                ws.cell(row, headers["data_source"], "current_poc_catalog_146")


def _model_label(variant_id: str, *, single: bool = False) -> str:
    base = VARIANT_LABELS.get(variant_id, variant_id)
    return f"{base}，单智能体" if single else base


def _platform_data_row(
    score: dict,
    *,
    label: str,
    data_nature: str,
    source_note: str,
) -> list:
    return [
        label,
        data_nature,
        rate_display(score["risk_num"], score["risk_den"]),
        score.get(MACRO_SUCCESS_RATE, "-"),
        rate_display(score.get("coverage_num", score.get("task_num")), score.get("coverage_den", score.get("task_den"))),
        f'{score["duration_seconds"] / 60:.2f} min（三目标平均，完整任务减人工等待）',
        score["authorization_required_count"],
        source_note,
    ]


def _table10_source_note(score: dict, *, variant: str, single: bool = False) -> str:
    per_target = score.get("per_target") or []
    parts = [
        f"{item.get('target_id')}: 漏洞{item.get('risk_num')}/{item.get('risk_den')}"
        for item in per_target
    ]
    mode = "单智能体" if single else "多智能体"
    return (
        f"{TABLE10_TARGETS_NOTE}；{variant} {mode} 各目标单轮实测；"
        f"漏洞检出与表6同口径（execution.vulnerable∪findings，分母30并集）；"
        f"耗时取三目标算术平均；"
        f"分目标：{'；'.join(parts)}"
    )


def _pgpt_table10_source_note(score: dict) -> str:
    per_target = score.get("per_target") or []
    parts = [
        f"{item.get('target_id')}: 漏洞{len(item.get('risk_hit_poc_files') or [])}/"
        f"{len(item.get('positive_poc_files') or [])} ({item.get('run_id')})"
        for item in per_target
    ]
    present = score.get("targets_present") or pgpt_available_targets()
    if len(present) == len(TABLE10_TARGETS):
        scope = "三目标单轮实测"
    else:
        scope = f"已测 {len(present)}/{len(TABLE10_TARGETS)} 目标"
    return (
        f"{TABLE10_TARGETS_NOTE}；PentestGPT GLM-5 {scope}；"
        f"漏洞检出与表6/10同口径（去重分母30）；耗时取已测目标算术平均；"
        f"分目标：{'；'.join(parts)}"
    )


def load_pgpt_table10_score(legacy_scores: list[dict]) -> tuple[dict, str]:
    aggregated = score_pentestgpt_report_set("PGPT_GLM5")
    if aggregated:
        present = aggregated.get("targets_present") or []
        if len(present) == len(TABLE10_TARGETS):
            nature = "三目标单轮实测（对照基线）"
        else:
            nature = f"部分目标实测（{len(present)}/{len(TABLE10_TARGETS)}，对照基线）"
        return aggregated, nature
    if len(legacy_scores) != 1:
        raise SystemExit(
            f"PentestGPT 三目标结果未齐且 legacy 结果数量异常: {len(legacy_scores)}"
        )
    legacy = legacy_scores[0]
    return {
        "completed_poc_count": int(legacy["completed_poc_count"]),
        "poc_selection_num": int(
            legacy.get("correct_selected_count")
            or round(float(legacy["poc_selection_recall"]) * int(legacy["expected_poc_count"]))
        ),
        "poc_selection_den": int(legacy["expected_poc_count"]),
        "risk_num": int(
            legacy.get("true_finding_count")
            or round(float(legacy["risk_recall"]) * int(legacy["positive_poc_count"]))
        ),
        "risk_den": int(legacy["positive_poc_count"]),
        "coverage_num": int(
            legacy.get("completed_task_count")
            or round(float(legacy["task_completion_rate"]) * int(legacy["task_count"]))
        ),
        "coverage_den": int(legacy["task_count"]),
        "task_num": int(
            legacy.get("completed_task_count")
            or round(float(legacy["task_completion_rate"]) * int(legacy["task_count"]))
        ),
        "task_den": int(legacy["task_count"]),
        "duration_seconds": float(legacy["task_duration_excluding_manual_wait_seconds"] or 0),
        "authorization_required_count": int(legacy["authorization_required_count"]),
        "per_target": [{
            "target_id": "REAL-CAR",
            "run_id": legacy.get("run_id"),
            "risk_hit_poc_files": [],
            "positive_poc_files": [],
        }],
        "targets_present": ["REAL-CAR"],
        "run_ids": [legacy.get("run_id")],
    }, "REAL-CAR 单轮实测（对照基线）"


def platform_comparison_rows(pgpt: dict, *, pgpt_data_nature: str) -> list[list]:
    write_manifest_cache()
    platform = score_platform_report_set("ZHIPU_MULTI")
    gpt_multi = score_platform_report_set("GPT_MULTI")
    gpt_single = score_platform_report_set("GPT_SINGLE")
    platform["model"] = _model_label("ZHIPU")
    gpt_multi["model"] = _model_label("GPT")
    gpt_single["model"] = _model_label("GPT", single=True)
    return [
        [
            "系统/模型",
            "数据性质",
            RECALL_AT_GT,
            MACRO_SUCCESS_RATE,
            COVERAGE,
            AVG_LATENCY,
            "高风险授权触发数",
            "数据来源与口径",
        ],
        _platform_data_row(
            platform,
            label=f'本文平台（{platform["model"]}）',
            data_nature="三目标单轮实测（多智能体）",
            source_note=_table10_source_note(platform, variant="GLM-5"),
        ),
        _platform_data_row(
            gpt_multi,
            label=f'本文平台（{gpt_multi["model"]}）',
            data_nature="三目标单轮实测（多智能体）",
            source_note=_table10_source_note(gpt_multi, variant="GPT-5.4-mini"),
        ),
        _platform_data_row(
            gpt_single,
            label=f'本文平台（{gpt_single["model"]}）',
            data_nature="三目标单轮实测（单智能体）",
            source_note=_table10_source_note(gpt_single, variant="GPT-5.4-mini", single=True),
        ),
        _platform_data_row(
            pgpt,
            label="PentestGPT（GLM-5）",
            data_nature=pgpt_data_nature,
            source_note=_pgpt_table10_source_note(pgpt),
        ),
    ]


def update_table10_platform_sheet(
    wb,
    pgpt: dict,
    *,
    pgpt_data_nature: str,
    index: int | None = None,
) -> None:
    rows = platform_comparison_rows(pgpt, pgpt_data_nature=pgpt_data_nature)
    sheet_index = index
    if sheet_index is None and TABLE10_PLATFORM_SHEET in wb.sheetnames:
        sheet_index = wb.sheetnames.index(TABLE10_PLATFORM_SHEET)
    replace_sheet(
        wb,
        TABLE10_PLATFORM_SHEET,
        rows,
        TABLE10_PLATFORM_WIDTHS,
        index=sheet_index,
    )


def update_paper_books_table10(pgpt: dict, *, pgpt_data_nature: str) -> list[str]:
    updated: list[str] = []
    for book in PAPER_BOOKS:
        if not book.is_file():
            continue
        wb = load_workbook(book)
        update_table10_platform_sheet(wb, pgpt, pgpt_data_nature=pgpt_data_nature)
        wb.save(book)
        updated.append(str(book))
    return updated


def update_total_book(pgpt: dict, *, pgpt_data_nature: str) -> None:
    wb = load_workbook(TOTAL_BOOK)
    rename_if_present(wb, "表10_反思重入汇总", "反思重入_汇总")
    rename_if_present(wb, "表10_三目标明细", "反思重入_三目标明细")
    update_pass_columns(wb)
    index = wb.sheetnames.index("反思重入_汇总") if "反思重入_汇总" in wb.sheetnames else 0
    update_table10_platform_sheet(wb, pgpt, pgpt_data_nature=pgpt_data_nature, index=index)
    note = wb["论文关键指标"]
    metric_row = None
    for row in range(2, note.max_row + 1):
        if note.cell(row, 1).value == "当前 PoC 库规模":
            metric_row = row
            break
    metric_row = metric_row or note.max_row + 1
    note.cell(metric_row, 1, "当前 PoC 库规模")
    note.cell(metric_row, 2, 146)
    note.cell(metric_row, 3, "新增 PoC 未倒算进既有130项冻结实验集")
    wb.save(TOTAL_BOOK)


def update_summary_book(pgpt: dict, *, pgpt_data_nature: str) -> None:
    wb = load_workbook(SUMMARY_BOOK)
    rename_if_present(wb, "表10_反思重入", "反思重入实验")
    update_pass_columns(wb)
    update_library_coverage(wb["表4_PoC攻击面覆盖"])

    # 表7/表8 只更新数据单元格，保留论文表格原有列结构与样式。
    paper_book = Path("/Users/queen/Desktop/ICV_POC_research/论文/论文内表格（新）.xlsx")
    if workbook_has_strict_tables(paper_book):
        update_paper_workbook(paper_book)
        restore_summary_tables_from_paper_book(paper_book)
        wb = load_workbook(SUMMARY_BOOK)
    elif TABLE7_SHEET in wb.sheetnames and TABLE8_SHEET in wb.sheetnames:
        update_paper_workbook(SUMMARY_BOOK)

    reflection_sheet = "反思重入实验" if "反思重入实验" in wb.sheetnames else "表10_反思重入"
    if reflection_sheet not in wb.sheetnames:
        reflection_sheet = "表10_反思重入汇总" if "表10_反思重入汇总" in wb.sheetnames else reflection_sheet
    t7_total = next((r for r in strict_rows("table7_total_by_model_group") if r.get("组别") == "D"), {})
    reflection_rows = [
        ["指标", "数值", "数据来源"],
        ["首次执行失败步骤数", 8, "三目标Agent报告聚合"],
        ["触发反思次数", 14, "三目标Agent报告聚合"],
        ["定向重跑次数", 14, "三目标Agent报告聚合"],
        ["补充侦察/补证问题数", 20, "三目标Agent报告聚合"],
        ["补证成功数量", t7_total.get("Agent命中阳性数", ""), "比较结果聚合"],
        ["最终 Recall@GT", t7_total.get(RECALL_AT_GT, ""), t7_total.get(HITS_AT_GT_FRACTION, "")],
    ]
    reflection_index = wb.sheetnames.index(reflection_sheet) if reflection_sheet in wb.sheetnames else len(wb.sheetnames)
    replace_sheet(wb, reflection_sheet, reflection_rows, [24, 15, 30], index=reflection_index)

    insert_at = wb.sheetnames.index(reflection_sheet) if reflection_sheet in wb.sheetnames else len(wb.sheetnames)
    update_table10_platform_sheet(wb, pgpt, pgpt_data_nature=pgpt_data_nature, index=insert_at)
    wb.save(SUMMARY_BOOK)


def save_table10_aggregate_json() -> None:
    write_manifest_cache()
    payload = {
        "targets": list(TABLE10_TARGETS),
        "targets_note": TABLE10_TARGETS_NOTE,
        "aggregation": (
            "Primary=Recall@GT (30-union |GT|); auxiliary=Coverage & Macro Success Rate (SR); "
            "Avg. Latency = mean net wall-clock across three targets"
        ),
        "hit_rule": "execution.vulnerable=True ∪ report.findings",
        "rows": {
            set_id: score_platform_report_set(set_id)
            for set_id in ("ZHIPU_MULTI", "GPT_MULTI", "GPT_SINGLE")
        },
    }
    out = STRICT_DATA / "table10_platform_three_targets.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def save_pgpt_aggregate_json(pgpt_score: dict) -> None:
    out = STRICT_DATA / "table10_pentestgpt_three_targets.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(pgpt_score, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    legacy_scores = json.loads(PGPT_RESULTS.read_text(encoding="utf-8"))
    pgpt_score, pgpt_data_nature = load_pgpt_table10_score(legacy_scores)
    save_table10_aggregate_json()
    save_pgpt_aggregate_json(pgpt_score)
    update_total_book(pgpt_score, pgpt_data_nature=pgpt_data_nature)
    update_summary_book(pgpt_score, pgpt_data_nature=pgpt_data_nature)
    paper_updated = update_paper_books_table10(pgpt_score, pgpt_data_nature=pgpt_data_nature)
    print(
        json.dumps(
            {
                "updated": [str(TOTAL_BOOK), str(SUMMARY_BOOK), *paper_updated],
                "pgpt_targets": pgpt_score.get("targets_present"),
                "pgpt_data_nature": pgpt_data_nature,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
