#!/usr/bin/env python3
"""Build a safe onboarding inventory for public PoC/EXP references.

This tool does not ingest exploit logic. It normalizes spreadsheet entries,
checks current repository coverage, and classifies whether a CVE is a good
candidate for safe plugin onboarding.
"""
from __future__ import annotations

import argparse
import csv
import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SERVER_DIR = Path(__file__).resolve().parent
POCS_DIR = SERVER_DIR / "pocs"
REPORTS_DIR = SERVER_DIR / "reports"


@dataclass
class PublicPocRecord:
    index: int
    cve_id: str
    year: str
    category: str
    product: str
    component: str
    vuln_type: str
    summary: str
    public_status: str
    public_source_url: str
    advisory_url: str
    automotive_relevance: str
    confidence: str
    repo_status: str
    onboarding_bucket: str
    target_domain: str
    plugin_mode: str


def _read_rows(xlsx_path: Path) -> Iterable[PublicPocRecord]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    repo_text = _repo_text()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        cve_id = str(row[1]).strip()
        category = str(row[3] or "").strip()
        product = str(row[4] or "").strip()
        component = str(row[5] or "").strip()
        vuln_type = str(row[6] or "").strip()
        summary = str(row[7] or "").strip()
        public_status = str(row[8] or "").strip()
        public_source_url = str(row[9] or "").strip()
        advisory_url = str(row[10] or "").strip()
        automotive_relevance = str(row[11] or "").strip()
        confidence = str(row[12] or "").strip()
        repo_status = "covered" if cve_id in repo_text else "missing"
        target_domain = _domain_for_entry(category, product, component, vuln_type)
        plugin_mode = _plugin_mode_for_entry(category, vuln_type, summary, public_status)
        onboarding_bucket = _bucket_for_entry(repo_status, plugin_mode, category, public_status)
        yield PublicPocRecord(
            index=int(row[0]),
            cve_id=cve_id,
            year=str(row[2] or "").strip(),
            category=category,
            product=product,
            component=component,
            vuln_type=vuln_type,
            summary=summary,
            public_status=public_status,
            public_source_url=public_source_url,
            advisory_url=advisory_url,
            automotive_relevance=automotive_relevance,
            confidence=confidence,
            repo_status=repo_status,
            onboarding_bucket=onboarding_bucket,
            target_domain=target_domain,
            plugin_mode=plugin_mode,
        )


def _repo_text() -> str:
    chunks: list[str] = []
    for path in POCS_DIR.rglob("*.py"):
        chunks.append(path.read_text(encoding="utf-8", errors="ignore"))
    return "\n".join(chunks)


def _domain_for_entry(category: str, product: str, component: str, vuln_type: str) -> str:
    blob = " ".join((category, product, component, vuln_type)).lower()
    if any(token in blob for token in ("wifi", "wpa", "802.11", "frag", "krack", "sae", "eap-pwd")):
        return "wireless"
    if "bluetooth" in blob or "hid" in blob or "l2cap" in blob or "bluez" in blob:
        return "wireless"
    if any(token in blob for token in ("android", "aaos", "webview", "chromium", "libwebp", "gstreamer", "stagefright")):
        return "application"
    if any(token in blob for token in ("tesla", "firmware", "ota", "updater", "odin")):
        return "network"
    return "advanced"


def _plugin_mode_for_entry(category: str, vuln_type: str, summary: str, public_status: str) -> str:
    blob = " ".join((category, vuln_type, summary, public_status)).lower()
    if any(token in blob for token in ("cache-based", "version", "signing", "key reuse", "information leak")):
        return "authenticated_check"
    if any(token in blob for token in ("wifi", "krack", "frag", "bluetooth", "hid", "injection")):
        return "active_protocol_validation"
    if any(token in blob for token in ("media", "webp", "gstreamer", "stagefright", "codec", "overflow")):
        return "local_parser_validation"
    if any(token in blob for token in ("firmware", "update", "odin", "token", "diagnostic")):
        return "workflow_validation"
    return "manual_review_only"


def _bucket_for_entry(repo_status: str, plugin_mode: str, category: str, public_status: str) -> str:
    if repo_status == "covered":
        return "already-covered"
    if plugin_mode in {"active_protocol_validation", "workflow_validation"}:
        return "priority-onboard"
    if plugin_mode == "local_parser_validation":
        return "lab-onboard"
    if "公开PoC" in public_status or "公开EXP" in public_status or "公开PoC/EXP" in public_status:
        return "review-for-safe-adaptation"
    if "android" in category.lower():
        return "authenticated-first"
    return "manual-triage"


def _write_csv(records: list[PublicPocRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(records[0]).keys()) if records else ["cve_id"])
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))


def _write_json(records: list[PublicPocRecord], path: Path) -> None:
    path.write_text(json.dumps([asdict(item) for item in records], ensure_ascii=False, indent=2), encoding="utf-8")


def _write_markdown(records: list[PublicPocRecord], path: Path) -> None:
    covered = [item for item in records if item.repo_status == "covered"]
    missing = [item for item in records if item.repo_status == "missing"]
    priority = [item for item in missing if item.onboarding_bucket == "priority-onboard"]
    lab = [item for item in missing if item.onboarding_bucket == "lab-onboard"]
    auth = [item for item in missing if item.onboarding_bucket == "authenticated-first"]

    lines = [
        "# Public PoC/EXP Onboarding Inventory",
        "",
        "This inventory normalizes spreadsheet entries into safe plugin-onboarding candidates.",
        "",
        "## Summary",
        "",
        f"- Total entries: `{len(records)}`",
        f"- Already covered in repo: `{len(covered)}`",
        f"- Missing from repo: `{len(missing)}`",
        f"- Priority onboard: `{len(priority)}`",
        f"- Lab-only onboard: `{len(lab)}`",
        f"- Authenticated-first: `{len(auth)}`",
        "",
        "## Priority Onboard",
        "",
    ]
    for item in priority:
        lines.append(f"- `{item.cve_id}` | {item.product} | {item.vuln_type} | {item.target_domain} | {item.plugin_mode}")
    lines.extend(["", "## Lab Onboard", ""])
    for item in lab:
        lines.append(f"- `{item.cve_id}` | {item.product} | {item.vuln_type} | {item.target_domain} | {item.plugin_mode}")
    lines.extend(["", "## Authenticated First", ""])
    for item in auth:
        lines.append(f"- `{item.cve_id}` | {item.product} | {item.vuln_type} | {item.target_domain} | {item.plugin_mode}")
    lines.extend([
        "",
        "## Safe Adaptation Rules",
        "",
        "1. Do not import public exploit payloads or weaponized trigger code directly.",
        "2. Convert entries into one of: authenticated check, active protocol validation, workflow validation, or local parser validation.",
        "3. Preserve public-source provenance in metadata and require explicit risk gates for any disruptive behavior.",
        "4. If a public artifact only supports lab reproduction, keep the plugin lab-scoped and do not present it as a general remote exploit check.",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a safe onboarding inventory for public PoC/EXP spreadsheets.")
    parser.add_argument("xlsx_path", help="Path to the public PoC/EXP spreadsheet")
    parser.add_argument("--prefix", default="public_poc_exp_inventory_20260708", help="Report filename prefix under server/reports")
    args = parser.parse_args()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    records = list(_read_rows(Path(args.xlsx_path)))
    json_path = REPORTS_DIR / f"{args.prefix}.json"
    csv_path = REPORTS_DIR / f"{args.prefix}.csv"
    md_path = REPORTS_DIR / f"{args.prefix}.md"
    _write_json(records, json_path)
    _write_csv(records, csv_path)
    _write_markdown(records, md_path)

    print(f"entries={len(records)}")
    print(f"covered={sum(1 for item in records if item.repo_status == 'covered')}")
    print(f"missing={sum(1 for item in records if item.repo_status == 'missing')}")
    print(f"json={json_path}")
    print(f"csv={csv_path}")
    print(f"md={md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
